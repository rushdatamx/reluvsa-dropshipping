"""
Generador de la plantilla de publicaciones masivas para Mercado Libre.

Convierte el catálogo de un proveedor en el .xlsx de 36 columnas que Gaby sube a
ML — el mismo formato de 'PENDIENTES ACDELCO.xlsx', que es la plantilla real que
ella usa hoy.

⭐ UNA PIEZA GENERA N PUBLICACIONES. Es lo que vuelve masivo el módulo: medido en
su archivo, 83 filas salieron de sólo 22 SKUs (~3.8x). Cada aplicación de la
columna "Aplicaciones Principales" es una publicación distinta; el SKU, el
precio, la descripción y las imágenes se repiten IDÉNTICOS y lo único que cambia
es el TÍTULO (el coche al que le sirve). Así compite por más búsquedas con la
misma pieza.

LA DESCRIPCIÓN es cabeza variable + cuerpo fijo, tal como ella lo pidió:
*"que pueda ponerse una sola descripción donde yo pueda poner mi descripción
base, pero lo que cambie sea el principio que es el tema de equivalencias o
compatibilidades"*. El cuerpo (garantía, horarios, facturación) se guarda una vez
por proveedor y no se vuelve a tocar.

⬜ LAS IMÁGENES VAN VACÍAS a propósito (confirmado por Gaby el 2026-08-19):
*"las subo a autozur, copio el link que me arroja y lo pongo, pero esto seguiría
siendo manual"*. El catálogo del proveedor no trae fotos. Ella las pega después
de generar el archivo.
"""
import re
from dataclasses import dataclass, field
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from services.aplicaciones_kg import Aplicacion, parse_aplicaciones
from services.perfiles_catalogo import PerfilCatalogo
from services.precio_publicacion import ParametrosPrecio, calcular_precio

# Las 36 columnas de la plantilla, en el orden EXACTO que espera Mercado Libre.
# 🔴 No reordenar ni renombrar: ML lee por posición y encabezado.
COLUMNAS = [
    "Titulo", "Categoria", "Precio", "Moneda(MXN,ARS,COP)", "Cantidad",
    "Tipo Publicacion (clasica,premium)", "Condición(nuevo,usado)",
    "Garantia(CANTIDAD [días, meses, años]_[FABRICA,VENDEDOR])",
    "Modo Envio(me1,me2)", "Dimensiones(ALTcm,ANCHcm,LONGcm,PESOgr)",
    "Envio Gratis(si,no)", "SKU", "Descripcion", "Tienda Oficial",
    "Imagen1", "Imagen2", "Imagen3", "Imagen4", "Imagen5",
    "Imagen6", "Imagen7", "Imagen8", "Imagen9", "Imagen10",
    "UPC", "Marca", "Talla", "Color", "Modelo",
    "Canal(mercadolibre,mshops,ambos)",
    "AG", "CAUPLAS", "KG", "KIM", "MATRIZ", "VAZLO",
]

# Las 6 bodegas del final: el stock va en la columna del proveedor, 0 en las demás.
BODEGAS = ["AG", "CAUPLAS", "KG", "KIM", "MATRIZ", "VAZLO"]

# Las constantes que Gaby marcó en AMARILLO PURO: siempre van iguales.
# "Envio Gratis" no pertenece aquí: se deriva del precio final de cada fila.
CONSTANTES = {
    "Moneda(MXN,ARS,COP)": "MXN",
    "Cantidad": 10,
    "Tipo Publicacion (clasica,premium)": "Clasica",
    "Condición(nuevo,usado)": "Nuevo",
    "Garantia(CANTIDAD [días, meses, años]_[FABRICA,VENDEDOR])": "2meses_vendedor",
    "Modo Envio(me1,me2)": "me2",
    "Canal(mercadolibre,mshops,ambos)": "mercadolibre",
}

# Tope real de Mercado Libre para el título de una publicación.
MAX_TITULO = 60


@dataclass
class ConfiguracionProveedor:
    """Lo que Gaby edita por proveedor, sin tocar código."""

    codigo_bodega: str
    descripcion_base: str = ""      # el cuerpo fijo (garantía, horarios, facturación)
    marca: str = ""
    categoria_ml: str = ""          # 'MLM163963' — la categoría de ML
    cantidad: int = 10
    params_precio: ParametrosPrecio = field(default_factory=ParametrosPrecio)


@dataclass
class FilaPublicacion:
    """Una fila del Excel = una publicación de ML."""

    titulo: str
    sku: str
    linea: str
    precio: Optional[float]
    descripcion: str
    aplicacion: str
    truncada: bool = False
    imagenes: List[str] = field(default_factory=list)
    fila_origen: Optional[int] = None


# Alias aprobados para las líneas reales del master. Sólo se usan si el nombre
# completo no cabe; modelo, cilindrada y años nunca se recortan.
ALIAS_PRODUCTO = {
    "banda de accesorios": "Banda Accesorios", "banda de tiempo": "Banda Tiempo",
    "bomba de agua": "Bomba Agua", "bomba de agua auxiliar": "Bomba Agua Aux",
    "deposito de anticongelante": "Deposito Anticongelante", "fan clutch": "Fan Clutch",
    "kit de banda de accesorios": "Kit Banda Accesorios",
    "kit de banda de distribucion": "Kit Banda Distribucion",
    "kit de banda de distribucion c/bomba": "Kit Banda Distrib c/Bomba",
    "kit de cadena de distribucion": "Kit Cadena Distribucion",
    "manguera moldeada": "Manguera Moldeada", "motoventilador": "Motoventilador",
    "polea de accesorios": "Polea Accesorios", "polea de distribucion": "Polea Distribucion",
    "sensor de temperatura": "Sensor Temperatura", "tapon de deposito": "Tapon Deposito",
    "tapon de radiador": "Tapon Radiador",
    "tensor hidraulico de distribucion": "Tensor Hidraulico Distrib",
    "tensor de accesorios": "Tensor Accesorios", "tensor de accesorios hd": "Tensor Accesorios HD",
    "tensor de accesorios hidraulico": "Tensor Accesorios Hidraulico",
    "toma de agua": "Toma Agua", "toma de agua housing": "Toma Agua Housing",
    "toma de agua con termostato": "Toma Agua c/Termostato",
    "tubo de enfriamiento": "Tubo Enfriamiento",
}

# Alias exclusivos del master CAUPLAS. Son la tercera y última alternativa:
# primero se intenta el nombre completo con armadora, luego sin armadora y sólo
# entonces se abrevia el producto. Nunca se recorta el vehículo ni los años.
ALIAS_PRODUCTO_CAUPLAS = {
    "deposito de refrigerante de motor": "Depósito Refrigerante",
    "manguera radiador": "Mang. Radiador",
    "manguera calefaccion": "Mang. Calef.",
    "manguera refrigeracion": "Mang. Refrig.",
    "manguera descarga de gases": "Mang. Descarga Gases",
    "manguera para frenos hidraulicos": "Mang. Freno Hidráulico",
    "manguera circulacion de aire": "Mang. Circulación Aire",
    "refrigeracion": "Refrig.",
}


def _sin_acentos(txt):
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", str(txt).lower())
                   if unicodedata.category(c) != "Mn")


def _cilindrada(motor):
    m = re.search(r"(\d[.,]\d)", str(motor or ""))
    return m.group(1).replace(",", ".") if m else ""


def _titulo_master(producto, compat, anios) -> Optional[str]:
    producto = " ".join(str(producto or "").split())
    armadora = " ".join(str(compat.get("armadora") or "").split())
    modelo = " ".join(str(compat.get("modelo") or "").split())
    motor = _cilindrada(compat.get("motor"))
    sufijo = " ".join(x for x in (modelo, motor, anios) if x)
    alias = ALIAS_PRODUCTO.get(_sin_acentos(producto), producto)
    intentos = [f"{producto} P/ {armadora} {sufijo}", f"{producto} P/ {sufijo}",
                f"{alias} P/ {sufijo}"]
    for titulo in intentos:
        titulo = re.sub(r"\s+", " ", titulo).strip()
        if len(titulo) <= MAX_TITULO: return titulo
    return None


def _variantes_master(producto, compat):
    inicio, fin = compat["inicio"], compat["fin"]
    if inicio == fin:
        titulo = _titulo_master(producto, compat, str(inicio))
        return ([titulo] if titulo else []), ([] if titulo else [(str(inicio), "Título excede 60 caracteres")])
    variantes, excluidas = [], []
    rango = f"{inicio}/{fin}"
    titulo = _titulo_master(producto, compat, rango)
    (variantes if titulo else excluidas).append(titulo or (rango, "Título excede 60 caracteres"))
    anios = list(range(inicio, fin + 1)); corte = (len(anios) + 1) // 2

    def agregar_bloque(bloque):
        texto = " ".join(map(str, bloque)); t = _titulo_master(producto, compat, texto)
        if t: variantes.append(t)
        elif len(bloque) > 1:
            mitad = (len(bloque) + 1) // 2
            agregar_bloque(bloque[:mitad]); agregar_bloque(bloque[mitad:])
        else: excluidas.append((texto, "Título excede 60 caracteres"))
    agregar_bloque(anios[:corte]); agregar_bloque(anios[corte:])
    return variantes, excluidas


def _titulo_cauplas(producto, compat, anios) -> Optional[str]:
    producto = " ".join(str(producto or "").split())
    armadora = " ".join(str(compat.get("armadora") or "").split())
    modelo = " ".join(str(compat.get("modelo") or "").split())
    cilindrada = " ".join(str(compat.get("motor") or "").split())
    if compat.get("universal"):
        sufijo = "Universal"
        intentos = [f"{producto} P/ {sufijo}"]
    else:
        sufijo = " ".join(x for x in (modelo, cilindrada, anios) if x)
        intentos = [f"{producto} P/ {armadora} {sufijo}", f"{producto} P/ {sufijo}"]
    alias = ALIAS_PRODUCTO_CAUPLAS.get(_sin_acentos(producto))
    if alias:
        intentos.append(f"{alias} P/ {sufijo}")
    for titulo in intentos:
        titulo = re.sub(r"\s+", " ", titulo).strip()
        if len(titulo) <= MAX_TITULO:
            return titulo
    return None


def _variantes_cauplas(producto, compat):
    if compat.get("universal"):
        titulo = _titulo_cauplas(producto, compat, "")
        return ([titulo] if titulo else []), ([] if titulo else [("All", "Título excede 60 caracteres")])
    inicio, fin = compat["inicio"], compat["fin"]
    if inicio == fin:
        titulo = _titulo_cauplas(producto, compat, str(inicio))
        return ([titulo] if titulo else []), ([] if titulo else [(str(inicio), "Título excede 60 caracteres")])
    variantes, excluidas = [], []
    rango = f"{inicio}/{fin}"
    titulo = _titulo_cauplas(producto, compat, rango)
    (variantes if titulo else excluidas).append(titulo or (rango, "Título excede 60 caracteres"))
    anios = list(range(inicio, fin + 1))
    corte = (len(anios) + 1) // 2

    def agregar_bloque(bloque):
        texto = " ".join(map(str, bloque))
        titulo_bloque = _titulo_cauplas(producto, compat, texto)
        if titulo_bloque:
            variantes.append(titulo_bloque)
        elif len(bloque) > 1:
            mitad = (len(bloque) + 1) // 2
            agregar_bloque(bloque[:mitad])
            agregar_bloque(bloque[mitad:])
        else:
            excluidas.append((texto, "Título excede 60 caracteres"))

    agregar_bloque(anios[:corte])
    agregar_bloque(anios[corte:])
    return variantes, excluidas


def _titular(nombre_pieza: str, app: Aplicacion) -> str:
    """Arma el título como lo escribe Gaby: 'Pieza P/ Coche Motor Años'.

    Formato copiado de su plantilla real:
        'Sensor Map De Aire Admision P/ Truck H200 2.5 2016 2017 2018'
    Se recorta a 60 caracteres (tope de ML) quitando años del final, nunca a
    media palabra: un título cortado se ve como error en la publicación.
    """
    pieza = " ".join(w.capitalize() for w in str(nombre_pieza or "").split())
    vehiculo = " ".join(w.capitalize() if w.isupper() and len(w) > 3 else w
                        for w in app.vehiculo.split())
    # Del motor sólo interesan los litros ('L4 1.6L' -> '1.6').
    litros = ""
    if app.motor:
        m = re.search(r"(\d[.,]\d)", app.motor)
        if m:
            litros = m.group(1).replace(",", ".")

    cabeza = f"{pieza} P/ {vehiculo}".strip()
    if litros:
        cabeza = f"{cabeza} {litros}"

    anios = app.anios.split()
    titulo = f"{cabeza} {' '.join(anios)}".strip()
    # Si no cabe, se van soltando años del final (el rango sigue siendo válido).
    while len(titulo) > MAX_TITULO and anios:
        anios.pop()
        titulo = f"{cabeza} {' '.join(anios)}".strip()
    return titulo[:MAX_TITULO].strip()


def _describir(nombre_pieza: str, clave: str, apps: List[Aplicacion],
               config: ConfiguracionProveedor) -> str:
    """Cabeza variable (pieza + OEM + compatibilidades) + cuerpo fijo.

    Las compatibilidades listan TODAS las aplicaciones de la pieza, no sólo la de
    esta fila: el comprador quiere ver si le sirve a su coche.
    """
    partes = [str(nombre_pieza or "").strip(), "", "OEM:", str(clave or "").strip()]

    utiles = [a for a in apps if not a.truncada]
    if utiles:
        partes += ["", "Compatibilidades:"]
        for a in utiles:
            linea = " ".join(x for x in (a.vehiculo, a.motor, a.anios) if x)
            partes.append(linea)

    if config.descripcion_base:
        partes += ["", config.descripcion_base.strip()]

    return "\n".join(partes)


def _describir_master(pieza, config):
    partes = [pieza.get("producto", "")]
    if pieza.get("oems"):
        partes += ["", "OEM: " + ", ".join(pieza["oems"])]
    guias = [c.get("guia") for c in pieza.get("compatibilidades", []) if c.get("guia")]
    if guias: partes += ["", "Compatibilidades:"] + list(dict.fromkeys(guias))
    if pieza.get("especificaciones"):
        partes += ["", "Especificaciones:"] + pieza["especificaciones"]
    if pieza.get("caracteristicas"):
        partes += ["", "Características:"] + pieza["caracteristicas"]
    if config.descripcion_base: partes += ["", config.descripcion_base.strip()]
    return "\n".join(str(x).strip() for x in partes)


def _describir_cauplas(pieza, config):
    partes = ["Productos y especificaciones:"]
    for par in pieza.get("productos_especificaciones", []):
        producto, especificacion = par.get("producto", ""), par.get("especificacion", "")
        partes.append(f"{producto} — {especificacion}" if especificacion else producto)
    if pieza.get("oems"):
        partes += ["", "OEM: " + " | ".join(pieza["oems"])]
    equivalencias = [(marca, codigos) for marca, codigos in pieza.get("equivalencias", {}).items()
                     if codigos]
    if equivalencias:
        partes += ["", "Equivalencias:"]
        partes += [f"{marca}: {' | '.join(codigos)}" for marca, codigos in equivalencias]
    productos = [x.get("producto", "") for x in pieza.get("productos_especificaciones", [])]
    if any("manguera" in _sin_acentos(producto) for producto in productos):
        medidas = [(nombre, valores) for nombre, valores in pieza.get("medidas", {}).items() if valores]
        if medidas:
            partes += ["", "Medidas:"]
            partes += [f"{nombre}: {' | '.join(valores)}" for nombre, valores in medidas]
    compatibilidades = []
    for compat in pieza.get("compatibilidades", []):
        if compat.get("universal"):
            texto = "Universal"
        else:
            anios = (str(compat["inicio"]) if compat["inicio"] == compat["fin"]
                     else f"{compat['inicio']}-{compat['fin']}")
            texto = " ".join(x for x in (compat.get("armadora"), compat.get("modelo"),
                                         compat.get("motor"), anios) if x)
        if _sin_acentos(texto) not in {_sin_acentos(x) for x in compatibilidades}:
            compatibilidades.append(texto)
    if compatibilidades:
        partes += ["", "Compatibilidades:"] + compatibilidades
    if config.descripcion_base:
        partes += ["", config.descripcion_base.strip()]
    return "\n".join(str(x).strip() for x in partes)


def generar_filas_con_reporte(piezas, config):
    filas, exclusiones, deduplicadas, vistos = [], [], 0, set()
    for pieza in piezas:
        if pieza.get("formato") == "master_cauplas":
            descripcion = _describir_cauplas(pieza, config)
            for compat in pieza.get("compatibilidades", []):
                producto = compat.get("producto", "")
                precio = calcular_precio(pieza.get("costo"), producto, config.params_precio)
                titulos, fuera = _variantes_cauplas(producto, compat)
                for anios, motivo in fuera:
                    exclusiones.append({"fila": compat.get("fila"), "clave": pieza.get("clave"),
                        "armadora": compat.get("armadora"), "modelo": compat.get("modelo"),
                        "anio": anios, "inicio": compat.get("inicio"), "fin": compat.get("fin"),
                        "motivo": motivo})
                for titulo in titulos:
                    par = (str(pieza.get("clave")).upper(), " ".join(_sin_acentos(titulo).split()))
                    if par in vistos:
                        deduplicadas += 1
                        continue
                    vistos.add(par)
                    filas.append(FilaPublicacion(
                        titulo, str(pieza.get("clave") or "").strip(), producto, precio,
                        descripcion, "Universal" if compat.get("universal") else
                        f"{compat.get('inicio')}-{compat.get('fin')}", False, [], compat.get("fila"),
                    ))
            continue
        if pieza.get("formato") != "master_kg":
            for f in generar_filas([pieza], config):
                par = (f.sku.upper(), " ".join(_sin_acentos(f.titulo).split()))
                if par in vistos: deduplicadas += 1
                else: vistos.add(par); filas.append(f)
            continue
        precio = calcular_precio(pieza.get("costo"), pieza.get("linea", ""), config.params_precio)
        descripcion = _describir_master(pieza, config)
        for compat in pieza.get("compatibilidades", []):
            titulos, fuera = _variantes_master(pieza.get("producto"), compat)
            for anios, motivo in fuera:
                exclusiones.append({"fila": compat.get("fila"), "clave": pieza.get("clave"),
                    "armadora": compat.get("armadora"), "modelo": compat.get("modelo"),
                    "anio": anios, "inicio": compat.get("inicio"), "fin": compat.get("fin"), "motivo": motivo})
            for titulo in titulos:
                par = (str(pieza.get("clave")).upper(), " ".join(_sin_acentos(titulo).split()))
                if par in vistos: deduplicadas += 1; continue
                vistos.add(par)
                filas.append(FilaPublicacion(titulo, str(pieza.get("clave") or "").strip(),
                    pieza.get("producto", ""), precio, descripcion, compat.get("guia", ""), False,
                    pieza.get("imagenes", []), compat.get("fila")))
    return filas, {"exclusiones": exclusiones, "deduplicadas": deduplicadas}


def generar_filas(piezas, config: ConfiguracionProveedor,
                  incluir_truncadas: bool = False) -> List[FilaPublicacion]:
    """Expande cada pieza del catálogo en sus N publicaciones."""
    if piezas and piezas[0].get("formato") in {"master_kg", "master_cauplas"}:
        return generar_filas_con_reporte(piezas, config)[0]
    filas: List[FilaPublicacion] = []

    for pieza in piezas:
        res = parse_aplicaciones(pieza.get("aplicaciones"))
        apps = res.aplicaciones if incluir_truncadas else res.utiles
        if not apps:
            continue

        precio = calcular_precio(pieza.get("costo"), pieza.get("linea", ""),
                                 config.params_precio)
        descripcion = _describir(pieza.get("linea"), pieza.get("clave"),
                                 res.aplicaciones, config)

        for app in apps:
            filas.append(FilaPublicacion(
                titulo=_titular(pieza.get("linea"), app),
                sku=str(pieza.get("clave") or "").strip(),
                linea=str(pieza.get("linea") or "").strip(),
                precio=precio,
                descripcion=descripcion,
                aplicacion=app.texto,
                truncada=app.truncada,
            ))

    return filas


def escribir_xlsx(filas: List[FilaPublicacion], config: ConfiguracionProveedor,
                  destino) -> int:
    """Escribe el .xlsx con las 36 columnas listo para subir a ML."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Publicaciones"

    encabezado = Font(bold=True)
    relleno = PatternFill("solid", fgColor="FFFF00")
    for c, nombre in enumerate(COLUMNAS, 1):
        celda = ws.cell(1, c, nombre)
        celda.font = encabezado
        if nombre in CONSTANTES:
            celda.fill = relleno

    idx = {nombre: i + 1 for i, nombre in enumerate(COLUMNAS)}
    bodega = config.codigo_bodega.strip().upper()

    for f, fila in enumerate(filas, start=2):
        ws.cell(f, idx["Titulo"], fila.titulo)
        ws.cell(f, idx["Categoria"], config.categoria_ml)
        # Precio None deja la celda VACÍA (no 0): un 0 se publicaría como precio real.
        if fila.precio is not None:
            ws.cell(f, idx["Precio"], fila.precio)
            ws.cell(f, idx["Envio Gratis(si,no)"], "Si" if fila.precio >= 299.00 else "No")
        ws.cell(f, idx["SKU"], fila.sku)
        ws.cell(f, idx["Descripcion"], fila.descripcion).alignment = Alignment(wrap_text=True)
        ws.cell(f, idx["Marca"], config.marca)
        ws.cell(f, idx["Modelo"], fila.sku)
        for numero, imagen in enumerate(fila.imagenes[:5], 1):
            if imagen: ws.cell(f, idx[f"Imagen{numero}"], imagen)

        for nombre, valor in CONSTANTES.items():
            ws.cell(f, idx[nombre], valor)
        ws.cell(f, idx["Cantidad"], config.cantidad)

        # Stock en la bodega del proveedor, 0 en las demás.
        for b in BODEGAS:
            ws.cell(f, idx[b], config.cantidad if b == bodega else 0)

        # Imagen6..10 siempre quedan vacías. El master nuevo llena Imagen1..5.

    ws.freeze_panes = "A2"
    for col, ancho in (("A", 58), ("M", 60), ("L", 18)):
        ws.column_dimensions[col].width = ancho

    wb.save(destino)
    return len(filas)
