"""Lectura de catálogos de Publicaciones Masivas y cruce con ML.

Los masters CAUPLAS y KG tienen lectores independientes. CAUPLAS se detecta
primero por encabezados para que su estructura nunca caiga en el parser KG.
"""
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Dict, List, Set, Tuple

import openpyxl

from services.perfiles_catalogo import PerfilCatalogo

COL_ATT_SELLER_SKU, COL_TITULO = 16, 1
MASTER_HOJA = "BD_Catalogo"
MASTER_HEADERS = ("Armadora", "Modelo", "Motor", "Producto", "Año", "Inicio", "Fin",
                  "Clave", "Especificaciones", "Características", "Guía de Compradores",
                  "OEM", "Imagen 1", "Imagen 2", "Imagen 3", "Imagen 4", "Imagen 5")
MASTER_ANCLAS = ("Armadora", "Modelo", "Producto", "Clave", "Inicio", "Fin")
MASTER_HEADERS_COSTO = ("Gran Mayoreo", "Precio")
STOCK_HEADER = "stock"

CAUPLAS_HEADERS = (
    "armadora", "modelo", "cilindrada", "uso", "especificaciones", "fecha",
    "inicio", "fin", "cauplas", "alto", "largo", "diametro1", "diametro2",
    "continental", "dayco", "gates", "keepongreen", "meisterzats", "tepeyac",
    "oe", "Precio",
)
CAUPLAS_ANCLAS = ("armadora", "modelo", "uso", "fecha", "cauplas", "oe", "Precio")
CAUPLAS_EQUIVALENCIAS = (
    ("Continental", "continental"), ("Dayco", "dayco"), ("Gates", "gates"),
    ("KeepOnGreen", "keepongreen"), ("Meisterzats", "meisterzats"),
    ("Tepeyac", "tepeyac"),
)


def _texto(val) -> str:
    if val is None: return ""
    if isinstance(val, float) and val.is_integer(): return str(int(val))
    return re.sub(r"\s+", " ", str(val).strip())


def _normalizar(val) -> str:
    txt = unicodedata.normalize("NFD", _texto(val).casefold())
    return " ".join("".join(c for c in txt if unicodedata.category(c) != "Mn").split())


def normalizar_titulo(val) -> str:
    return _normalizar(val)


@dataclass
class ResultadoCatalogo:
    piezas: List[Dict]
    formato: str
    filas_master: int
    compatibilidades_validas: int
    compatibilidades_invalidas: int = 0
    duplicados_descartados: int = 0
    precio_presente: bool = True
    sku_sin_precio: int = 0
    sku_precio_inconsistente: int = 0
    errores: List[Dict] = field(default_factory=list)
    sku_unicos_master: int = 0
    universales: int = 0


def _error(fila, valores, motivo):
    return {"fila": fila, "clave": _texto(valores.get("clave")),
            "armadora": _texto(valores.get("armadora")), "modelo": _texto(valores.get("modelo")),
            "anio": _texto(valores.get("anio")), "inicio": _texto(valores.get("inicio")),
            "fin": _texto(valores.get("fin")), "motivo": motivo}


def _stock_entero(valor):
    """Valida el inventario que llega desde Excel sin convertir decimales.

    openpyxl entrega 10.0 como ``float``; ése sí representa el entero 10.
    En cambio 10.5, texto y celdas vacías no pueden convertirse en inventario.
    """
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None, "Stock vacío"
    if isinstance(valor, bool):
        return None, "Stock debe ser un entero mayor o igual a cero"
    if isinstance(valor, int):
        numero = valor
    elif isinstance(valor, float):
        if not valor.is_integer():
            return None, "Stock debe ser un entero mayor o igual a cero"
        numero = int(valor)
    else:
        texto = _texto(valor)
        if not re.fullmatch(r"[+-]?\d+", texto):
            return None, "Stock debe ser un entero mayor o igual a cero"
        numero = int(texto)
    if numero < 0:
        return None, "Stock debe ser un entero mayor o igual a cero"
    return numero, None


def _stocks_por_sku(registros, crear_error):
    """Devuelve ``SKU -> stock`` y deja fuera cualquier SKU inconsistente."""
    stocks, excluidos = {}, set()
    for sku, filas in registros.items():
        invalidas = [r for r in filas if r["motivo"]]
        if invalidas:
            excluidos.add(sku)
            for registro in invalidas:
                crear_error(registro, registro["motivo"])
            continue
        stock = filas[0]["stock"]
        distinta = next((r for r in filas[1:] if r["stock"] != stock), None)
        if distinta:
            excluidos.add(sku)
            crear_error(distinta, "Stock diferente dentro del SKU")
            continue
        stocks[sku] = stock
    return stocks, excluidos


def _anio_entero(valor):
    if isinstance(valor, bool): return None
    if isinstance(valor, int): return valor
    if isinstance(valor, float) and valor.is_integer(): return int(valor)
    txt = _texto(valor)
    return int(txt) if re.fullmatch(r"\d{4}", txt) else None


def _producto_canonico(valor) -> str:
    conectores = {"de", "del", "con", "c/bomba"}
    return " ".join(p.lower() if p.lower() in conectores else p.capitalize()
                    for p in _texto(valor).split())


def _headers_master(ws):
    valores = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    mapa = {_normalizar(v): i for i, v in enumerate(valores) if _texto(v)}
    encabezado_costo = next((h for h in MASTER_HEADERS_COSTO if _normalizar(h) in mapa), None)
    return all(_normalizar(h) in mapa for h in MASTER_HEADERS), mapa, encabezado_costo


def _parece_master(ws) -> bool:
    """Distingue un master incompleto de un catálogo legado.

    Cuatro de las seis columnas de identidad son suficientes para no interpretar
    accidentalmente Modelo como la línea de producto del formato anterior.
    """
    _, columnas, _ = _headers_master(ws)
    return sum(_normalizar(h) in columnas for h in MASTER_ANCLAS) >= 4


def _hoja_master(wb):
    # El nombre canónico sigue ganando cuando existe. La estructura permite que
    # el proveedor renombre la hoja sin cambiar el formato del archivo.
    canonica = wb[MASTER_HOJA] if MASTER_HOJA in wb.sheetnames else None
    if canonica is not None and _headers_master(canonica)[0]:
        return canonica
    completas = [ws for ws in wb.worksheets if _headers_master(ws)[0]]
    if completas:
        return completas[0]
    if canonica is not None:
        return canonica
    incompletas = [ws for ws in wb.worksheets if _parece_master(ws)]
    return incompletas[0] if incompletas else None


def _headers_cauplas(ws):
    valores = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    mapa = {_normalizar(v): i for i, v in enumerate(valores) if _texto(v)}
    return all(_normalizar(h) in mapa for h in CAUPLAS_HEADERS), mapa


def _parece_cauplas(ws) -> bool:
    _, columnas = _headers_cauplas(ws)
    return sum(_normalizar(h) in columnas for h in CAUPLAS_ANCLAS) >= 5


def _hoja_cauplas(wb):
    completas = [ws for ws in wb.worksheets if _headers_cauplas(ws)[0]]
    if completas:
        return completas[0]
    incompletas = [ws for ws in wb.worksheets if _parece_cauplas(ws)]
    return incompletas[0] if incompletas else None


def _partes_codigos(valor):
    """Separa listas del proveedor sin inventar estructura dentro del código."""
    return [p.strip() for p in re.split(r"\s*\|\s*", _texto(valor)) if p.strip() and p.strip() != "-"]


def _agregar_unico(lista, valor):
    if valor and _normalizar(valor) not in {_normalizar(x) for x in lista}:
        lista.append(valor)


def _error_cauplas(fila, sku, datos, motivo):
    return {
        "fila": fila, "clave": sku, "armadora": _texto(datos.get("armadora")),
        "modelo": _texto(datos.get("modelo")), "anio": _texto(datos.get("fecha")),
        "inicio": _texto(datos.get("inicio")), "fin": _texto(datos.get("fin")),
        "motivo": motivo,
    }


def _validar_anios_cauplas(datos):
    fecha = _texto(datos.get("fecha"))
    if fecha.casefold() == "all":
        return True, None, None, True, ""
    inicio, fin = _anio_entero(datos.get("inicio")), _anio_entero(datos.get("fin"))
    if inicio is None or fin is None or not (1900 <= inicio <= 2100 and 1900 <= fin <= 2100):
        return False, inicio, fin, False, "Inicio y Fin deben ser años enteros entre 1900 y 2100"
    if inicio > fin:
        return False, inicio, fin, False, "Inicio es mayor que Fin"
    anios_fecha = [int(x) for x in re.findall(r"(?<!\d)(?:19|20)\d{2}(?!\d)", fecha)]
    if not anios_fecha or min(anios_fecha) != inicio or max(anios_fecha) != fin:
        return False, inicio, fin, False, "Fecha no coincide con Inicio y Fin"
    return True, inicio, fin, False, ""


def _leer_master_cauplas(ws) -> ResultadoCatalogo:
    valido, columnas = _headers_cauplas(ws)
    if not valido or STOCK_HEADER not in columnas:
        faltan = [h for h in CAUPLAS_HEADERS if _normalizar(h) not in columnas]
        if STOCK_HEADER not in columnas:
            faltan.append("stock")
        raise ValueError(
            f"La hoja «{ws.title}» parece ser el master CAUPLAS, pero le faltan "
            "los encabezados requeridos: " + ", ".join(faltan)
        )

    def val(fila, nombre):
        i = columnas.get(_normalizar(nombre))
        return fila[i] if i is not None and i < len(fila) else None

    grupos, errores, sku_observados, stocks_registrados = {}, [], set(), {}
    filas_master = validas = universales = duplicados = 0
    max_col = max(columnas.values()) + 1
    for numero, fila in enumerate(ws.iter_rows(min_row=2, max_col=max_col, values_only=True), 2):
        if not any(_texto(x) for x in fila):
            continue
        filas_master += 1
        datos = {nombre: val(fila, encabezado) for nombre, encabezado in (
            ("armadora", "armadora"), ("modelo", "modelo"), ("cilindrada", "cilindrada"),
            ("producto", "uso"), ("especificacion", "especificaciones"), ("fecha", "fecha"),
            ("inicio", "inicio"), ("fin", "fin"), ("sku", "cauplas"), ("oem", "oe"),
        )}
        sku = _texto(datos["sku"])
        if sku:
            sku_observados.add(sku.upper())
            stock, motivo_stock = _stock_entero(val(fila, STOCK_HEADER))
            stocks_registrados.setdefault(sku.upper(), []).append({
                "fila": numero, "stock": stock, "motivo": motivo_stock, "datos": datos,
            })
        if not sku:
            errores.append(_error_cauplas(numero, sku, datos, "SKU CAUPLAS vacío"))
            continue
        ok, inicio, fin, universal, motivo = _validar_anios_cauplas(datos)
        if not ok:
            errores.append(_error_cauplas(numero, sku, datos, motivo))
            continue
        if universal:
            universales += 1
        else:
            validas += 1

        producto = _producto_canonico(datos["producto"])
        especificacion = _texto(datos["especificacion"])
        armadora, modelo, cilindrada = (_texto(datos[x]) for x in ("armadora", "modelo", "cilindrada"))
        g = grupos.setdefault(sku.upper(), {
            "clave": sku, "linea": producto, "lineas": [], "producto": producto,
            "formato": "master_cauplas", "compatibilidades": [],
            "productos_especificaciones": [], "oems": [],
            "equivalencias": {marca: [] for marca, _ in CAUPLAS_EQUIVALENCIAS},
            "medidas": {"Alto": [], "Largo": [], "Diámetro 1": [], "Diámetro 2": []},
            "imagenes": [], "_costos": set(), "filas_origen": [],
        })
        g["filas_origen"].append(numero)
        _agregar_unico(g["lineas"], producto)
        par_producto = {"producto": producto, "especificacion": especificacion}
        firma_producto = (_normalizar(producto), _normalizar(especificacion))
        if firma_producto not in {(_normalizar(x["producto"]), _normalizar(x["especificacion"]))
                                  for x in g["productos_especificaciones"]}:
            g["productos_especificaciones"].append(par_producto)

        compat = {"producto": producto, "armadora": armadora, "modelo": modelo,
                  "motor": cilindrada, "inicio": inicio, "fin": fin,
                  "universal": universal, "fila": numero}
        firma_compat = (_normalizar(producto), _normalizar(armadora), _normalizar(modelo),
                        _normalizar(cilindrada), inicio, fin, universal)
        firmas = {(_normalizar(c["producto"]), _normalizar(c["armadora"]),
                   _normalizar(c["modelo"]), _normalizar(c["motor"]), c["inicio"],
                   c["fin"], c["universal"]) for c in g["compatibilidades"]}
        if firma_compat in firmas:
            duplicados += 1
        else:
            g["compatibilidades"].append(compat)

        for codigo in _partes_codigos(datos["oem"]):
            _agregar_unico(g["oems"], codigo)
        for marca, encabezado in CAUPLAS_EQUIVALENCIAS:
            for codigo in _partes_codigos(val(fila, encabezado)):
                _agregar_unico(g["equivalencias"][marca], codigo)
        for etiqueta, encabezado in (("Alto", "alto"), ("Largo", "largo"),
                                     ("Diámetro 1", "diametro1"), ("Diámetro 2", "diametro2")):
            medida = _texto(val(fila, encabezado))
            try:
                es_cero = float(medida) == 0
            except (TypeError, ValueError):
                es_cero = False
            if medida and not es_cero:
                _agregar_unico(g["medidas"][etiqueta], medida)
        try:
            costo = float(val(fila, "Precio"))
            if costo > 0:
                g["_costos"].add(costo)
        except (TypeError, ValueError):
            pass

    def error_stock(registro, motivo):
        errores.append(_error_cauplas(registro["fila"], registro["datos"]["sku"],
                                      registro["datos"], motivo))

    stocks, sku_stock_invalido = _stocks_por_sku(stocks_registrados, error_stock)
    piezas = [g for clave, g in grupos.items() if clave not in sku_stock_invalido]
    sin_precio = inconsistentes = 0
    for g in piezas:
        g["stock"] = stocks[g["clave"].upper()]
        if len(g["_costos"]) == 1:
            g["costo"] = next(iter(g["_costos"]))
        else:
            g["costo"] = None
            if len(g["_costos"]) > 1:
                inconsistentes += 1
                errores.append(_error_cauplas(g["filas_origen"][0], g["clave"], {},
                                               "Precios diferentes dentro del SKU"))
            else:
                sin_precio += 1
        del g["_costos"]

    return ResultadoCatalogo(
        piezas, "master_cauplas", filas_master, validas,
        len([e for e in errores if e["motivo"] in {
            "Inicio y Fin deben ser años enteros entre 1900 y 2100",
            "Inicio es mayor que Fin", "Fecha no coincide con Inicio y Fin",
        }]), duplicados, True, sin_precio, inconsistentes, errores,
        len(sku_observados), universales,
    )


def _leer_master(ws) -> ResultadoCatalogo:
    valido, columnas, encabezado_costo = _headers_master(ws)
    if not valido or STOCK_HEADER not in columnas:
        faltan = [h for h in MASTER_HEADERS if _normalizar(h) not in columnas]
        if STOCK_HEADER not in columnas:
            faltan.append("stock")
        raise ValueError(
            f"La hoja «{ws.title}» parece ser el master KG, pero le faltan "
            "los encabezados requeridos: " + ", ".join(faltan)
        )
    costos_presentes = [h for h in MASTER_HEADERS_COSTO if _normalizar(h) in columnas]
    # Stock puede estar en cualquier posición; el costo no puede adelantarse a
    # las columnas estructurales (en particular Imagen 5), pero ya no se fija
    # a una letra absoluta.
    idx_imagen5 = columnas[_normalizar("Imagen 5")]
    if any(columnas[_normalizar(h)] < idx_imagen5 for h in costos_presentes):
        raise ValueError(
            "Gran Mayoreo o Precio debe ir después de Imagen 5; antes debía estar "
            "inmediatamente después de Imagen 5"
        )
    precio_presente = encabezado_costo is not None
    grupos, errores, vistos, claves_master, stocks_registrados = {}, [], set(), set(), {}
    filas_master = validas = duplicados = 0

    def val(fila, nombre):
        i = columnas.get(_normalizar(nombre))
        return fila[i] if i is not None and i < len(fila) else None

    for numero, fila in enumerate(ws.iter_rows(min_row=2, max_col=max(columnas.values()) + 1,
                                                values_only=True), 2):
        if not any(_texto(x) for x in fila): break
        filas_master += 1
        datos = {k: val(fila, n) for k, n in (("armadora", "Armadora"), ("modelo", "Modelo"),
                 ("motor", "Motor"), ("producto", "Producto"), ("anio", "Año"),
                 ("inicio", "Inicio"), ("fin", "Fin"), ("clave", "Clave"))}
        clave = _texto(datos["clave"]); inicio = _anio_entero(datos["inicio"]); fin = _anio_entero(datos["fin"])
        if clave:
            claves_master.add(clave.upper())
            stock, motivo_stock = _stock_entero(val(fila, STOCK_HEADER))
            stocks_registrados.setdefault(clave.upper(), []).append({
                "fila": numero, "stock": stock, "motivo": motivo_stock, "datos": datos,
            })
        if not clave:
            errores.append(_error(numero, datos, "Clave vacía")); continue
        if inicio is None or fin is None or not (1900 <= inicio <= 2100 and 1900 <= fin <= 2100):
            errores.append(_error(numero, datos, "Inicio y Fin deben ser años enteros entre 1900 y 2100")); continue
        if inicio > fin:
            errores.append(_error(numero, datos, "Inicio es mayor que Fin")); continue
        armadora, modelo, motor = (_texto(datos[x]) for x in ("armadora", "modelo", "motor"))
        producto = _producto_canonico(datos["producto"]); guia = _texto(val(fila, "Guía de Compradores"))
        firma = tuple(_normalizar(x) for x in fila)
        if firma in vistos: duplicados += 1; continue
        vistos.add(firma)
        g = grupos.setdefault(clave.upper(), {"clave": clave, "linea": producto, "producto": producto,
            "formato": "master_kg", "compatibilidades": [], "oems": [], "especificaciones": [],
            "caracteristicas": [], "imagenes": ["", "", "", "", ""], "_imagenes_sets": set(),
            "_costos": set(), "filas_origen": []})
        g["filas_origen"].append(numero)
        firma_compat = (_normalizar(armadora), _normalizar(modelo), _normalizar(motor), inicio, fin,
                        _normalizar(guia))
        existentes = {(_normalizar(c["armadora"]), _normalizar(c["modelo"]),
                       _normalizar(c["motor"]), c["inicio"], c["fin"], _normalizar(c["guia"]))
                      for c in g["compatibilidades"]}
        if firma_compat not in existentes:
            validas += 1
            g["compatibilidades"].append({"armadora": armadora, "modelo": modelo, "motor": motor,
                                          "inicio": inicio, "fin": fin, "guia": guia, "fila": numero})
        else:
            duplicados += 1
        for nombre, destino in (("OEM", "oems"), ("Especificaciones", "especificaciones"),
                                ("Características", "caracteristicas")):
            texto = _texto(val(fila, nombre))
            if texto and texto != "-" and _normalizar(texto) not in {_normalizar(x) for x in g[destino]}:
                g[destino].append(texto)
        imagenes = tuple(_texto(val(fila, "Imagen %d" % i)) for i in range(1, 6))
        if any(imagenes):
            g["_imagenes_sets"].add(imagenes)
            if not any(g["imagenes"]): g["imagenes"] = list(imagenes)
        if precio_presente:
            try:
                costo = float(val(fila, encabezado_costo))
                if costo > 0: g["_costos"].add(costo)
            except (TypeError, ValueError): pass

    def error_stock(registro, motivo):
        errores.append(_error(registro["fila"], registro["datos"], motivo))

    stocks, sku_stock_invalido = _stocks_por_sku(stocks_registrados, error_stock)
    piezas = [g for clave, g in grupos.items() if clave not in sku_stock_invalido]
    inconsistentes = sin_precio = 0
    for g in piezas:
        g["stock"] = stocks[g["clave"].upper()]
        if len(g["_costos"]) == 1: g["costo"] = next(iter(g["_costos"]))
        else:
            g["costo"] = None
            if len(g["_costos"]) > 1:
                inconsistentes += 1
                errores.append(_error(g["filas_origen"][0], {"clave": g["clave"]},
                                      "Precios diferentes dentro del SKU"))
            else: sin_precio += 1
        if len(g["_imagenes_sets"]) > 1:
            errores.append(_error(g["filas_origen"][0], {"clave": g["clave"]},
                                  "Imágenes inconsistentes dentro del SKU"))
        del g["_costos"], g["_imagenes_sets"]
    invalidas = sum(1 for e in errores if "Inicio" in e["motivo"])
    return ResultadoCatalogo(piezas, "master_kg", filas_master, validas, invalidas,
                             duplicados, precio_presente, sin_precio, inconsistentes, errores,
                             len(claves_master))


def _leer_legado(wb, perfil):
    ws = wb[perfil.nombre_hoja] if perfil.nombre_hoja else wb.worksheets[0]
    encabezados = [c.value for c in next(ws.iter_rows(min_row=perfil.fila_header,
                                                       max_row=perfil.fila_header))]
    columnas = {_normalizar(valor): i for i, valor in enumerate(encabezados) if _texto(valor)}
    if STOCK_HEADER not in columnas:
        raise ValueError(f"La hoja «{ws.title}» no tiene el encabezado requerido: stock")
    piezas, errores, stocks_registrados = [], [], {}
    for numero, fila in enumerate(ws.iter_rows(min_row=perfil.fila_header + 1, values_only=True),
                                  perfil.fila_header + 1):
        clave = _texto(fila[perfil.col_clave]) if len(fila) > perfil.col_clave else ""
        if not clave: continue
        celda = lambda i: fila[i] if i is not None and len(fila) > i else None
        stock, motivo_stock = _stock_entero(celda(columnas[STOCK_HEADER]))
        datos = {"clave": clave}
        stocks_registrados.setdefault(clave.upper(), []).append({
            "fila": numero, "stock": stock, "motivo": motivo_stock, "datos": datos,
        })
        piezas.append({"clave": clave, "linea": _texto(celda(perfil.col_linea)),
            "aplicaciones": celda(perfil.col_aplicaciones), "costo": celda(perfil.col_precio_costo),
            "codigo_barras": _texto(celda(perfil.col_codigo_barras)),
            "precio_publico": celda(perfil.col_precio_publico), "formato": "legado"})
    filas_master = len(piezas)
    stocks, sku_stock_invalido = _stocks_por_sku(
        stocks_registrados,
        lambda registro, motivo: errores.append(_error(registro["fila"], registro["datos"], motivo)),
    )
    piezas = [p for p in piezas if p["clave"].upper() not in sku_stock_invalido]
    for pieza in piezas:
        pieza["stock"] = stocks[pieza["clave"].upper()]
    return ResultadoCatalogo(piezas, "legado", filas_master, len(piezas), errores=errores)


def leer_catalogo_detallado(ruta, perfil: PerfilCatalogo) -> ResultadoCatalogo:
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    try:
        hoja_cauplas = _hoja_cauplas(wb)
        if hoja_cauplas is not None:
            return _leer_master_cauplas(hoja_cauplas)
        if perfil.codigo_bodega == "CAUPLAS":
            raise ValueError("El archivo no tiene la estructura esperada del master CAUPLAS")
        hoja_master = _hoja_master(wb)
        return _leer_master(hoja_master) if hoja_master is not None else _leer_legado(wb, perfil)
    finally: wb.close()


def leer_catalogo(ruta, perfil): return leer_catalogo_detallado(ruta, perfil).piezas


def leer_publicaciones(ruta) -> Set[Tuple[str, str]]:
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    try:
        pares = set()
        for fila in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
            if len(fila) <= COL_ATT_SELLER_SKU: continue
            titulo = normalizar_titulo(fila[COL_TITULO] if len(fila) > COL_TITULO else "")
            for parte in _texto(fila[COL_ATT_SELLER_SKU]).split("&"):
                if parte.strip() and titulo: pares.add((parte.strip().upper(), titulo))
        return pares
    finally: wb.close()


def leer_skus_publicados(ruta): return {sku for sku, _ in leer_publicaciones(ruta)}


def cruzar_variantes(filas, publicados):
    pendientes, existentes, vistos, deduplicadas = [], [], set(), 0
    for fila in filas:
        par = (fila.sku.upper(), normalizar_titulo(fila.titulo))
        if par in vistos: deduplicadas += 1; continue
        vistos.add(par); (existentes if par in publicados else pendientes).append(fila)
    return {"pendientes": pendientes, "existentes": existentes, "deduplicadas": deduplicadas}


def cruzar(piezas: List[Dict], publicados: Set[str]) -> Dict:
    faltantes, ya = [], []
    for p in piezas: (ya if p["clave"].upper() in publicados else faltantes).append(p)
    return {"total_catalogo": len(piezas), "ya_publicadas": len(ya), "faltantes": len(faltantes),
            "piezas_faltantes": faltantes, "piezas_publicadas": ya}
