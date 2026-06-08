"""
Parser de 'Detalle envios de colecta.xlsx'.

Headers en fila 8 (index 7). 13 columnas con datos. Col J=10 'Lugar indicado',
col K=11 'Lugar desde donde hiciste el envío' (la que importa, según Gaby).
Col L=12 'Cumplió con lo indicado' (Sí/No).
"""
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from rapidfuzz import fuzz

from database import get_db

SHEETS_PRIORIDAD = ["Últimas 4 semanas", "Última semana"]

# Cruce envío -> venta (regla de Gaby). ML asigna a veces 2 folios distintos a
# la misma venta, así que num_venta no cruza fiable; cruzamos por fecha + título.
CRUCE_VENTANA_SEGUNDOS = 300       # ±5 min de tolerancia en la fecha de venta
CRUCE_TITULO_MIN = 85             # umbral fuzzy token_set_ratio sobre el título

# Mapeo Lugar (columna K) → codigo_bodega de proveedor
LUGAR_A_BODEGA = {
    "AG": "AG",
    "CAUPLAS": "CAUPLAS",
    "KG": "KG",
    "KIM": "KIM",
    "VAZLO": "VAZLO",
    "MATRIZ": None,  # bodega propia, no es proveedor dropshipping
}

MESES_ES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def _parse_fecha_es(val) -> Optional[datetime]:
    """Convierte 'viernes 8 may 2026 - 07:34 hs' o variaciones a datetime."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip().lower()
    m = re.search(r"(\d{1,2})\s+([a-z]{3})\s+(\d{4})(?:\s*-\s*(\d{1,2}):(\d{2}))?", s)
    if not m:
        return None
    day, mes_abv, year, hh, mm = m.groups()
    month = MESES_ES.get(mes_abv[:3])
    if not month:
        return None
    try:
        return datetime(int(year), month, int(day), int(hh or 0), int(mm or 0))
    except Exception:
        return None


def _find_header_idx(rows) -> Optional[int]:
    for i, row in enumerate(rows[:15]):
        if row and row[0] and "Fecha de la venta" in str(row[0]):
            return i
    return None


def _resolver_proveedor(conn, lugar: Optional[str]) -> Optional[int]:
    if not lugar:
        return None
    lugar_clean = lugar.strip().upper()
    codigo = LUGAR_A_BODEGA.get(lugar_clean)
    if codigo is None:
        return None
    row = conn.execute(
        "SELECT id FROM proveedores WHERE codigo_bodega = ?", (codigo,)
    ).fetchone()
    return row["id"] if row else None


def _to_dt(val) -> Optional[datetime]:
    """Normaliza un valor de fecha (datetime o ISO string de la BD) a datetime."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.fromisoformat(str(val))
    except (ValueError, TypeError):
        return None


def resolver_cruce_ventas(conn) -> dict:
    """Resuelve envios_colecta.num_venta_ml para todos los envíos cargados.

    Regla de Gaby: el # de venta NO es fiable como llave (ML usa 2 folios para una
    misma venta). Estrategia por envío:
      1. Directo: si su num_venta existe tal cual en ventas_ml -> confianza 1.0.
      2. Fallback: candidatas con fecha de venta dentro de ±5 min Y título fuzzy
         (token_set_ratio) >= 85. Se elige la de mayor score; desempate por menor
         diferencia de tiempo. Si num_venta coincide con alguna candidata, esa gana.
         Confianza = score/100; si hay 2+ candidatas >=85 con distinto num_venta,
         se marca como ambiguo restando 0.2 a la confianza (queda < umbral 0.6 de
         error) para que Gaby lo revise.
    Recalcula desde cero en cada corrida (idempotente).
    """
    ventas = conn.execute(
        "SELECT num_venta, fecha_venta, titulo FROM ventas_ml WHERE titulo IS NOT NULL"
    ).fetchall()
    ventas_por_num = {v["num_venta"]: v for v in ventas}
    # Pre-parseamos fechas de ventas una sola vez.
    ventas_dt = [(v["num_venta"], _to_dt(v["fecha_venta"]), v["titulo"] or "") for v in ventas]

    envios = conn.execute(
        "SELECT num_envio, num_venta, fecha_venta, titulo FROM envios_colecta"
    ).fetchall()

    cruces_directo = 0
    cruces_fuzzy = 0
    cruces_ambiguos = 0
    sin_cruce = 0

    for e in envios:
        num_venta = e["num_venta"]
        num_venta_ml = None
        confianza = None

        # 1) Directo por num_venta
        if num_venta and num_venta in ventas_por_num:
            num_venta_ml = num_venta
            confianza = 1.0
            cruces_directo += 1
        else:
            # 2) Fallback fecha (±ventana) + título fuzzy
            e_dt = _to_dt(e["fecha_venta"])
            e_tit = e["titulo"] or ""
            if e_dt and e_tit:
                candidatas = []
                for v_num, v_dt, v_tit in ventas_dt:
                    if not v_dt:
                        continue
                    dt = abs((v_dt - e_dt).total_seconds())
                    if dt <= CRUCE_VENTANA_SEGUNDOS:
                        sc = fuzz.token_set_ratio(e_tit, v_tit)
                        if sc >= CRUCE_TITULO_MIN:
                            candidatas.append((sc, dt, v_num))
                if candidatas:
                    # mayor score, luego menor diferencia de tiempo
                    candidatas.sort(key=lambda x: (-x[0], x[1]))
                    # desempate: si el num_venta del envío está entre las candidatas, gana
                    elegido = candidatas[0]
                    for c in candidatas:
                        if num_venta and c[2] == num_venta:
                            elegido = c
                            break
                    nums_distintos = {c[2] for c in candidatas}
                    ambiguo = len(nums_distintos) > 1
                    num_venta_ml = elegido[2]
                    confianza = round(elegido[0] / 100.0, 3)
                    if ambiguo:
                        confianza = round(confianza - 0.2, 3)
                        cruces_ambiguos += 1
                    cruces_fuzzy += 1
                else:
                    sin_cruce += 1
            else:
                sin_cruce += 1

        conn.execute(
            "UPDATE envios_colecta SET num_venta_ml=?, match_cruce_confianza=? WHERE num_envio=?",
            (num_venta_ml, confianza, e["num_envio"]),
        )

    return {
        "cruces_directo": cruces_directo,
        "cruces_fecha_titulo": cruces_fuzzy,
        "cruces_ambiguos": cruces_ambiguos,
        "envios_sin_cruce": sin_cruce,
        "total_cruzados": cruces_directo + cruces_fuzzy,
    }


def parse_colecta(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = next((s for s in SHEETS_PRIORIDAD if s in wb.sheetnames), wb.sheetnames[0])
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = _find_header_idx(rows)
    if header_idx is None:
        raise ValueError("No se detectó fila de encabezados en detalle de colecta")

    inserted = 0
    updated = 0
    sin_proveedor = 0

    with get_db() as conn:
        for row in rows[header_idx + 1:]:
            if not row or len(row) < 13 or not row[1]:
                continue

            num_envio = str(row[1]).strip()
            num_venta = str(row[2]).strip() if row[2] else None
            lugar_indicado = str(row[9]).strip() if row[9] else None
            lugar_real_raw = str(row[10]).strip() if row[10] else None
            lugar_real = (
                None if (lugar_real_raw and "sin información" in lugar_real_raw.lower())
                else lugar_real_raw
            )
            cumplio = row[11]
            cumplio_sla = 1 if (cumplio and str(cumplio).strip().lower() in ("sí", "si")) else (0 if cumplio else None)
            excluido_raw = row[12] if len(row) > 12 else None
            excluido = 1 if (excluido_raw and str(excluido_raw).strip().lower() in ("sí", "si")) else 0

            proveedor_id = _resolver_proveedor(conn, lugar_real)
            if proveedor_id is None and lugar_real:
                sin_proveedor += 1

            data = (
                num_envio,
                num_venta,
                _parse_fecha_es(row[0]),
                str(row[4]).strip() if row[4] else None,  # título
                str(row[5]).strip() if row[5] else None,  # tiempo max
                str(row[6]).strip() if row[6] else None,  # tiempo real
                lugar_indicado,
                lugar_real,
                proveedor_id,
                cumplio_sla,
                excluido,
            )

            existing = conn.execute(
                "SELECT num_envio, lugar_override FROM envios_colecta WHERE num_envio = ?", (num_envio,)
            ).fetchone()

            if existing:
                # No pisamos lugar_override; si ya hay override, recalculamos proveedor desde el override
                override = existing["lugar_override"]
                if override:
                    proveedor_id = _resolver_proveedor(conn, override) or proveedor_id
                conn.execute(
                    """UPDATE envios_colecta SET num_venta=?, fecha_venta=?, titulo=?, tiempo_max_envio=?,
                                                 tiempo_real_envio=?, lugar_indicado=?, lugar_real=?,
                                                 proveedor_id=?, cumplio_sla=?, excluido_analisis=?
                       WHERE num_envio=?""",
                    (data[1], data[2], data[3], data[4], data[5], data[6], data[7],
                     proveedor_id, data[9], data[10], num_envio),
                )
                updated += 1
            else:
                conn.execute(
                    """INSERT INTO envios_colecta
                       (num_envio, num_venta, fecha_venta, titulo, tiempo_max_envio, tiempo_real_envio,
                        lugar_indicado, lugar_real, proveedor_id, cumplio_sla, excluido_analisis)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    data,
                )
                inserted += 1

        # Resolver el cruce envío -> venta (fecha + título) ya con todo cargado.
        cruce = resolver_cruce_ventas(conn)

    return {
        "ok": True,
        "sheet_used": sheet_name,
        "inserted": inserted,
        "updated": updated,
        "envios_sin_proveedor_inferido": sin_proveedor,
        **cruce,
    }
