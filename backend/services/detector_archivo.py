"""
Detector del tipo de Excel subido, por su CONTENIDO (no por el nombre, que se
puede renombrar). Candado para que Gaby no suba el archivo equivocado en la
sección equivocada (ej. el Excel de colecta en el botón de Ventas).

Huellas únicas verificadas con los archivos reales:
- Ventas ML : hoja "Ventas MX" o header con "# de venta" + "Depósito".
- Colecta   : hojas "Última semana"/"Últimas 4 semanas", A1 "Envíos con colecta",
              o header con "Fecha de la venta" + "# de envío".
- Albarán   : header con "# de venta" + "albarán", SIN las anclas de ventas/colecta.
              (Excel simple de 2 columnas que arma Gaby: venta -> # de albarán.)
- Kits      : header con "componente" + "cantidad" + "paquete"/"kit" en la 1a hoja.
              (Excel de relación kit -> componentes: Paquete / Componente / Cantidad.
              NO por nombre de hoja: el control interno de Gaby tiene una hoja "KITS".)
Son mutuamente excluyentes: un reporte de ventas nunca trae "Envíos con colecta"
ni la hoja de colecta, el de albaranes es el único que trae "albarán" sin las
otras anclas, y el de kits es el único que trae "componente" + "cantidad".
"""
from pathlib import Path
from typing import Optional

import openpyxl


# Firmas de hoja (vienen de los parsers: SHEETS_PRIORIDAD y "Ventas MX").
_HOJAS_COLECTA = {"última semana", "últimas 4 semanas"}
_HOJA_VENTAS = "ventas mx"


def _texto_primeras_filas(ws, max_filas: int = 12) -> str:
    """Concatena en minúsculas las primeras celdas de las primeras filas, para
    buscar textos ancla sin importar la columna exacta."""
    partes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_filas:
            break
        for c in row:
            if c is not None and str(c).strip():
                partes.append(str(c).strip().lower())
    return " | ".join(partes)


def detectar_tipo_xlsx(path: Path) -> Optional[str]:
    """Devuelve 'ventas_ml', 'colecta', 'albaran', 'kits' o None según el contenido.

    None = no se reconoce como ninguno (archivo ajeno o corrupto).
    Nunca lanza: ante cualquier error de lectura, devuelve None.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None

    try:
        hojas = {s.strip().lower() for s in wb.sheetnames}

        # --- Señales por nombre de hoja (lo más barato y específico) ---
        if _HOJA_VENTAS in hojas:
            return "ventas_ml"
        if hojas & _HOJAS_COLECTA:
            return "colecta"

        # --- Señales por contenido de las primeras filas de la 1a hoja ---
        ws = wb[wb.sheetnames[0]]
        texto = _texto_primeras_filas(ws)

        # Kits: el header trae "componente" + "cantidad" + "paquete"/"kit". Va antes
        # de albarán: es el único con "componente", y no comparte anclas con ventas/colecta.
        # NOTA: deliberadamente NO usamos señal por nombre de hoja "KITS" — el workbook
        # de control interno de Gaby tiene 40+ hojas (una llamada "KITS") y daría falso
        # positivo. La detección por header de la 1a hoja es específica del Excel real
        # de relación kits (cuya 1a hoja ES la de Paquete/Componente/Cantidad).
        if "componente" in texto and "cantidad" in texto and ("paquete" in texto or "kit" in texto):
            return "kits"

        # Colecta: ancla muy específica del encabezado del reporte ML de colecta.
        if "envíos con colecta" in texto or ("fecha de la venta" in texto and "# de envío" in texto):
            return "colecta"

        # Ventas ML: el header trae "# de venta" + "depósito" juntos.
        if "# de venta" in texto and "depósito" in texto:
            return "ventas_ml"

        # Albarán: Excel simple de Gaby con "# de venta" (o "venta") + "albarán/albaran".
        # Va después de ventas/colecta para que esas anclas más específicas ganen primero
        # (un reporte de ventas/colecta nunca trae la palabra "albarán").
        tiene_venta = "# de venta" in texto or "venta" in texto
        tiene_albaran = "albarán" in texto or "albaran" in texto
        if tiene_venta and tiene_albaran:
            return "albaran"

        return None
    finally:
        wb.close()
