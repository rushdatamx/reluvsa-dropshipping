"""
Parser del reporte de Ventas Mercado Libre.

El archivo tiene 66 columnas con 2 niveles de header (fila índice 4 = categorías
mergeadas tipo 'Ventas'/'Publicaciones', fila índice 5 = nombres reales). Datos
arrancan en fila índice 6. El header se detecta dinámicamente. Solo nos importan:
- # de venta, Depósito (bodega de origen), Fecha de venta, Estado, Unidades, Total
- SKU, Título, Comprador, Estado del comprador, Forma de entrega
- Factura adjunta, Unidades devueltas, Reclamos

La columna 'Depósito' marca la bodega (MATRIZ/KIM/CAUPLAS/VAZLO/...). MATRIZ es bodega
propia de RELUVSA (no dropshipping); el portal la oculta por defecto en Ventas.

Las fechas vienen en español largo: "13 de mayo de 2026 23:43".
"""
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from database import get_db

# Meses en español, nombre completo (formato del reporte de Ventas ML).
MESES_ES_LARGO = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11,
    "diciembre": 12,
}

COLS_INTERES = {
    "# de venta": "num_venta",
    "Fecha de venta": "fecha_venta",
    "Estado": "estado",
    "Unidades": "unidades",
    "Total (MXN)": "total",
    "SKU": "sku",
    "Título de la publicación": "titulo",
    "Comprador": "comprador",
    "Estado": "comprador_estado",  # cuidado: duplicado con "Estado" de venta
    "Forma de entrega": "forma_entrega",
    "Factura adjunta": "factura_adjunta_ml",
}


def _find_header_rows(rows):
    """Detecta la fila con > 10 celdas no vacías que tiene los nombres reales."""
    for i, row in enumerate(rows[:15]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty > 10:
            return i
    return None


def _parse_fecha(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    # Formato real de ML: "13 de mayo de 2026 23:43" (hora opcional).
    m = re.search(
        r"(\d{1,2})\s+de\s+([a-záéíóú]+)\s+de\s+(\d{4})(?:\s+(\d{1,2}):(\d{2}))?",
        s.lower(),
    )
    if m:
        day, mes_nom, year, hh, mm = m.groups()
        month = MESES_ES_LARGO.get(mes_nom)
        if month:
            try:
                return datetime(int(year), month, int(day), int(hh or 0), int(mm or 0))
            except Exception:
                pass
    # Fallback: ISO (por si algún día el reporte cambia de formato).
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def _to_int(val, default=None):
    """Cast defensivo a int. Las celdas reales traen espacios sueltos (' '),
    floats ('1.0') y vacíos que int() no tolera."""
    if val is None:
        return default
    s = str(val).strip()
    if not s:
        return default
    try:
        return int(float(s))  # tolera '1.0'
    except (ValueError, TypeError):
        return default


def _to_float(val, default=None):
    if val is None:
        return default
    s = str(val).strip()
    if not s:
        return default
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def parse_ventas_ml(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Ventas MX"] if "Ventas MX" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = _find_header_rows(rows)
    if header_idx is None:
        raise ValueError("No se detectó fila de encabezados en el reporte de ventas ML")

    cat_row = rows[header_idx - 1] if header_idx > 0 else [None] * len(rows[header_idx])
    name_row = rows[header_idx]

    col_map: dict[str, int] = {}
    for j, (cat, name) in enumerate(zip(cat_row, name_row)):
        if not name:
            continue
        cat_str = str(cat).strip() if cat else ""
        name_str = str(name).strip()
        key = f"{cat_str}|{name_str}".strip("|")
        col_map[key] = j
        col_map[name_str] = j  # fallback

    def col(*candidates) -> Optional[int]:
        for c in candidates:
            if c in col_map:
                return col_map[c]
        return None

    idx_num_venta = col("Ventas|# de venta", "# de venta")
    idx_deposito = col("Depósito", "Deposito")  # bodega de origen (MATRIZ/KIM/CAUPLAS/...)
    idx_fecha = col("Ventas|Fecha de venta", "Fecha de venta")
    idx_estado = col("Ventas|Estado")
    idx_unidades = col("Ventas|Unidades", "Unidades")
    idx_total = col("Ventas|Total (MXN)", "Total (MXN)")
    idx_sku = col("Publicaciones|SKU", "SKU")
    idx_titulo = col("Publicaciones|Título de la publicación", "Título de la publicación")
    idx_comprador = col("Compradores|Comprador", "Comprador")
    idx_comp_estado = col("Compradores|Estado")
    idx_forma_entrega = col("Envíos|Forma de entrega", "Forma de entrega")
    idx_factura_ml = col("Facturación al comprador|Factura adjunta", "Factura adjunta")
    idx_devolucion = col("Devoluciones|Unidades")
    idx_reclamo_abierto = col("Reclamos|Reclamo abierto")
    idx_reclamo_cerrado = col("Reclamos|Reclamo cerrado")

    if idx_num_venta is None or idx_sku is None:
        raise ValueError("Columnas obligatorias (# de venta, SKU) no encontradas")

    inserted = 0
    updated = 0
    skipped = 0

    with get_db() as conn:
        for row in rows[header_idx + 1:]:
            num_venta = row[idx_num_venta]
            if not num_venta:
                continue
            num_venta = str(num_venta).strip()
            if num_venta in {"# de venta", ""}:
                skipped += 1
                continue

            data = {
                "num_venta": num_venta,
                "sku": str(row[idx_sku]).strip() if idx_sku is not None and row[idx_sku] else None,
                "deposito": str(row[idx_deposito]).strip() if idx_deposito is not None and row[idx_deposito] else None,
                "fecha_venta": _parse_fecha(row[idx_fecha]) if idx_fecha is not None else None,
                "estado": str(row[idx_estado]).strip() if idx_estado is not None and row[idx_estado] else None,
                "titulo": str(row[idx_titulo]).strip() if idx_titulo is not None and row[idx_titulo] else None,
                "unidades": _to_int(row[idx_unidades]) if idx_unidades is not None else None,
                "total": _to_float(row[idx_total]) if idx_total is not None else None,
                "comprador": str(row[idx_comprador]).strip() if idx_comprador is not None and row[idx_comprador] else None,
                "comprador_estado": str(row[idx_comp_estado]).strip() if idx_comp_estado is not None and row[idx_comp_estado] else None,
                "forma_entrega": str(row[idx_forma_entrega]).strip() if idx_forma_entrega is not None and row[idx_forma_entrega] else None,
                "factura_adjunta_ml": str(row[idx_factura_ml]).strip() if idx_factura_ml is not None and row[idx_factura_ml] else None,
                "devolucion_unidades": _to_int(row[idx_devolucion], default=0) if idx_devolucion is not None else 0,
                "reclamo_abierto": 1 if idx_reclamo_abierto is not None and row[idx_reclamo_abierto] else 0,
                "reclamo_cerrado": 1 if idx_reclamo_cerrado is not None and row[idx_reclamo_cerrado] else 0,
            }

            existing = conn.execute(
                "SELECT num_venta FROM ventas_ml WHERE num_venta = ?", (num_venta,)
            ).fetchone()

            if existing:
                conn.execute(
                    """UPDATE ventas_ml SET sku=?, deposito=?, fecha_venta=?, estado=?, titulo=?, unidades=?, total=?,
                                            comprador=?, comprador_estado=?, forma_entrega=?, factura_adjunta_ml=?,
                                            devolucion_unidades=?, reclamo_abierto=?, reclamo_cerrado=?
                       WHERE num_venta=?""",
                    (
                        data["sku"], data["deposito"], data["fecha_venta"], data["estado"], data["titulo"], data["unidades"],
                        data["total"], data["comprador"], data["comprador_estado"], data["forma_entrega"],
                        data["factura_adjunta_ml"], data["devolucion_unidades"], data["reclamo_abierto"],
                        data["reclamo_cerrado"], num_venta,
                    ),
                )
                updated += 1
            else:
                conn.execute(
                    """INSERT INTO ventas_ml (num_venta, sku, deposito, fecha_venta, estado, titulo, unidades, total,
                                              comprador, comprador_estado, forma_entrega, factura_adjunta_ml,
                                              devolucion_unidades, reclamo_abierto, reclamo_cerrado)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        num_venta, data["sku"], data["deposito"], data["fecha_venta"], data["estado"], data["titulo"],
                        data["unidades"], data["total"], data["comprador"], data["comprador_estado"],
                        data["forma_entrega"], data["factura_adjunta_ml"], data["devolucion_unidades"],
                        data["reclamo_abierto"], data["reclamo_cerrado"],
                    ),
                )
                inserted += 1

        # Re-resolver el cruce envío -> venta: si la colecta se cargó antes que
        # las ventas, ahora ya hay ventas con qué cruzar (regla fecha + título).
        from services.parser_colecta import resolver_cruce_ventas
        cruce = resolver_cruce_ventas(conn)

    return {
        "ok": True,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        **cruce,
    }
