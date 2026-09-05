"""Regresión del master CAUPLAS en Publicaciones Masivas."""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from services.generador_plantilla import (
    COLUMNAS, ConfiguracionProveedor, generar_filas_con_reporte, escribir_xlsx,
)
from services.imagenes_cauplas import asignar_imagenes_cauplas, leer_imagenes_cauplas
from services.parser_catalogo import cruzar_variantes, leer_catalogo_detallado, leer_publicaciones
from services.perfiles_catalogo import perfil_de, proveedores_soportados

_ok = _fail = 0


def check(nombre, condicion, extra=""):
    global _ok, _fail
    if condicion:
        _ok += 1
        print(f"  ✓ {nombre}")
    else:
        _fail += 1
        print(f"  ✗ {nombre}  {extra}")


HEADERS = [
    "Auto/Camioneta", "armadora", "modelo", "motor", "cilindrada", "cilindros",
    "disposicion", "combustible", "uso", "especificaciones", "fecha", "inicio",
    "fin", "cauplas", "imagen", "alto", "largo", "diametro1", "diametro2", "apa",
    "continental", "dayco", "gates", "keepongreen", "meisterzats", "tepeyac", "oe",
    "compatibilidad", "estatus", "abc", "codigo SAT", "Precio", "  STOCK  ",
]


def fila(sku, producto="Manguera Radiador", especificacion="Superior", fecha="2010-2012",
         inicio=2010, fin=2012, precio=100, armadora="GENERAL MOTORS", modelo="SILVERADO",
         cilindrada="5.3L", medidas=(10, 40, 32, 32), stock=10):
    valores = [None] * len(HEADERS)
    datos = {
        "armadora": armadora, "modelo": modelo, "cilindrada": cilindrada,
        "uso": producto, "especificaciones": especificacion, "fecha": fecha,
        "inicio": inicio, "fin": fin, "cauplas": sku, "alto": medidas[0],
        "largo": medidas[1], "diametro1": medidas[2], "diametro2": medidas[3],
        "continental": "C-1 | C-2", "dayco": "D-1", "gates": "G-1",
        "keepongreen": "KG-1", "meisterzats": "M-1", "tepeyac": "T-1",
        "oe": "OEM-1 | OEM-2", "Precio": precio, "  STOCK  ": stock,
    }
    indices = {nombre: i for i, nombre in enumerate(HEADERS)}
    for nombre, valor in datos.items():
        valores[indices[nombre]] = valor
    return valores


def guardar(wb):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


print("\n=== 1. PERFIL Y DETECCIÓN ===")
check("CAUPLAS y KG son proveedores soportados", proveedores_soportados() == ["CAUPLAS", "KG"])
check("la marca predeterminada de CAUPLAS es editable desde CAUPLAS",
      perfil_de("CAUPLAS").marca_ml == "CAUPLAS")
check("CAUPLAS sólo autoriza fotos de ImageKit",
      perfil_de("CAUPLAS").dominios_imagenes == ("ik.imagekit.io",))

print("\n=== 1B. CSV DE IMÁGENES CAUPLAS ===")
csv_imagenes = """Name,URL
2522.jpg,https://ik.imagekit.io/cauplas/2522.jpg
2522-2.jpg,https://ik.imagekit.io/cauplas/2522-2.jpg
2522-3.jpg,https://ik.imagekit.io/cauplas/2522-3.jpg
7498.jpg,https://ik.imagekit.io/cauplas/7498.jpg
7498-2.jpg,https://ik.imagekit.io/cauplas/7498-2.jpg
SKU-2.jpg,https://ik.imagekit.io/cauplas/sku-guion.jpg
2522-4.jpg,https://ik.imagekit.io/cauplas/2522-2.jpg
2522-11.jpg,https://ik.imagekit.io/cauplas/fuera-de-rango.jpg
SIN-SKU.jpg,https://ik.imagekit.io/cauplas/sin-match.jpg
"""
galerias, resumen = leer_imagenes_cauplas(csv_imagenes.encode(), {"2522", "7498", "SKU-2"})
check("ordena portada y fotos posteriores por sufijo",
      galerias["2522"][:3] == [
          "https://ik.imagekit.io/cauplas/2522.jpg", "https://ik.imagekit.io/cauplas/2522-2.jpg",
          "https://ik.imagekit.io/cauplas/2522-3.jpg"], galerias["2522"])
check("7498 llena Imagen1 e Imagen2", galerias["7498"][:2] == [
      "https://ik.imagekit.io/cauplas/7498.jpg", "https://ik.imagekit.io/cauplas/7498-2.jpg"])
check("el SKU completo con guion numérico gana al sufijo",
      galerias["SKU-2"][0] == "https://ik.imagekit.io/cauplas/sku-guion.jpg" and not galerias["SKU-2"][1])
check("omite URLs repetidas, más de diez y nombres sin match",
      resumen["urls_detectadas"] == 9 and resumen["urls_validas"] == 6 and
      resumen["urls_omitidas"] == 3 and resumen["urls_sin_match"] == 1, resumen)
try:
    leer_imagenes_cauplas(b"SKU,URL\n2522,https://ik.imagekit.io/a.jpg\n", {"2522"})
    encabezados_invalidos = False
except ValueError:
    encabezados_invalidos = True
check("CSV sin encabezado Name se rechaza", encabezados_invalidos)

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Master renombrado"
ws.append([f"  {h.upper()}  " if h else h for h in HEADERS])
ws.append(fila(123.0)); ws.append(fila("00123", especificacion="Inferior", precio=150))
ws.append(fila("UNI-1", producto="Termostato", fecha="All", inicio=None, fin=None))
ws.append(fila("BAD-1", fecha="2011-2015", inicio=2011, fin=2012))
ws.append(fila("FORMULA", precio="=10+20"))
ruta = guardar(wb)
r = leer_catalogo_detallado(ruta, perfil_de("CAUPLAS"))
check("detecta CAUPLAS por encabezados normalizados", r.formato == "master_cauplas")
check("SKU numérico pierde .0 y SKU texto conserva ceros",
      {p["clave"] for p in r.piezas} >= {"123", "00123"}, {p["clave"] for p in r.piezas})
check("All se cuenta como universal", r.universales == 1)
check("K contradictorio con L-M se excluye con fila, SKU y motivo",
      r.compatibilidades_invalidas == 1 and r.errores[0]["clave"] == "BAD-1" and
      "no coincide" in r.errores[0]["motivo"], r.errores)
formula = next(p for p in r.piezas if p["clave"] == "FORMULA")
check("fórmula sin valor cacheado deja precio vacío, nunca cero", formula["costo"] is None)

incompleto = openpyxl.Workbook(); wi = incompleto.active
wi.append([h for h in HEADERS if h != "oe"])
try:
    leer_catalogo_detallado(guardar(incompleto), perfil_de("CAUPLAS"))
    mensaje = ""
except ValueError as exc:
    mensaje = str(exc)
check("master CAUPLAS incompleto se rechaza con encabezado faltante",
      "master CAUPLAS" in mensaje and "oe" in mensaje, mensaje)

sin_stock = openpyxl.Workbook(); sin_stock.active.append([h for h in HEADERS if h != "  STOCK  "])
try:
    leer_catalogo_detallado(guardar(sin_stock), perfil_de("CAUPLAS"))
    mensaje_stock = ""
except ValueError as exc:
    mensaje_stock = str(exc)
check("master CAUPLAS sin stock se rechaza antes de analizar", "stock" in mensaje_stock.lower(), mensaje_stock)

print("\n=== 2. CONSOLIDACIÓN Y DESCRIPCIÓN ===")
wc = openpyxl.Workbook(); sc = wc.active; sc.append(HEADERS)
sc.append(fila(500, especificacion="Superior", precio=100))
segunda = fila("500", especificacion="Inferior", fecha="2013", inicio=2013, fin=2013,
               precio=100, medidas=(0, 45, 0, 35))
segunda[20] = "C-2 | C-3"; segunda[26] = "OEM-2 | OEM-3"; sc.append(segunda)
sc.append(fila("NO-MANG", producto="Termostato", especificacion="88 grados",
               fecha="All", inicio=None, fin=None, medidas=(99, 99, 99, 99)))
sc.append(fila("INC", fecha="2010", inicio=2010, fin=2010, precio=100))
sc.append(fila("INC", fecha="2011", inicio=2011, fin=2011, precio=120))
rc = leer_catalogo_detallado(guardar(wc), perfil_de("CAUPLAS"))
cfg = ConfiguracionProveedor(codigo_bodega="CAUPLAS", marca="CAUPLAS",
                             descripcion_base="BASE RELUVSA")
filas, reporte = generar_filas_con_reporte(rc.piezas, cfg)
f500 = next(f for f in filas if f.sku == "500")
check("consolida combinaciones I/J distintas", "Superior" in f500.descripcion and "Inferior" in f500.descripcion)
check("OEM se separa con | y sin duplicados", "OEM: OEM-1 | OEM-2 | OEM-3" in f500.descripcion)
check("equivalencias quedan separadas por marca y sin duplicados",
      "Continental: C-1 | C-2 | C-3" in f500.descripcion and "Dayco: D-1" in f500.descripcion)
check("manguera añade cm a Alto/Largo y mm a cada diámetro",
      "Alto: 10 cm" in f500.descripcion and "Largo: 40 cm | 45 cm" in f500.descripcion and
      "Diámetro 1: 32 mm" in f500.descripcion and "Diámetro 2: 32 mm | 35 mm" in f500.descripcion,
      f500.descripcion)
fterm = next(f for f in filas if f.sku == "NO-MANG")
check("producto que no es manguera omite medidas", "Medidas:" not in fterm.descripcion)
check("costos contradictorios no se eligen arbitrariamente",
      rc.sku_precio_inconsistente == 1 and
      next(p for p in rc.piezas if p["clave"] == "INC")["costo"] is None)
check("la descripción base queda al final", f500.descripcion.endswith("BASE RELUVSA"))
check("All genera el título universal exacto", fterm.titulo == "Termostato P/ Universal", fterm.titulo)

print("\n=== 3. TÍTULOS, EXPANSIÓN Y DEDUPLICACIÓN ===")
wa = openpyxl.Workbook(); sa = wa.active; sa.append(HEADERS)
sa.append(fila("ALIAS", producto="Manguera Circulación De Aire", armadora="MERCEDES-BENZ",
               modelo="MODELO EXTRA LARGO", fecha="2010-2014", inicio=2010, fin=2014))
sa.append(fila("DUP", fecha="2020", inicio=2020, fin=2020))
sa.append(fila("DUP", fecha="2020", inicio=2020, fin=2020))
ra = leer_catalogo_detallado(guardar(wa), perfil_de("CAUPLAS"))
fa, repa = generar_filas_con_reporte(ra.piezas, cfg)
alias = [f.titulo for f in fa if f.sku == "ALIAS"]
check("usa sólo alias controlado cuando el nombre completo no cabe",
      any(t.startswith("Mang. Circulación Aire P/") for t in alias), alias)
check("genera rango y bloques cronológicos", any("2010/2014" in t for t in alias) and len(alias) >= 3, alias)
check("todos los títulos respetan estrictamente 60 caracteres", all(len(f.titulo) <= 60 for f in fa))
check("deduplica por SKU + título normalizado", sum(f.sku == "DUP" for f in fa) == 1)

print("\n=== 4. CRUCE Y PLANTILLA DE 36 COLUMNAS ===")
pub = openpyxl.Workbook(); ps = pub.active; ps.append([""] * 17)
row = [""] * 17; row[1] = f500.titulo.upper(); row[16] = "OTRO&500"; ps.append(row)
publicados = leer_publicaciones(guardar(pub)); cruce = cruzar_variantes(filas, publicados)
check("cruce SKU+título reconoce paquetes separados por &", len(cruce["existentes"]) == 1)

salida = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); salida.close()
galeria_500, _ = leer_imagenes_cauplas(
    b"Name,URL\n500.jpg,https://ik.imagekit.io/cauplas/500.jpg\n500-2.jpg,https://ik.imagekit.io/cauplas/500-2.jpg\n",
    {"500", "NO-MANG", "INC"},
)
asignar_imagenes_cauplas(filas, galeria_500)
check("las variantes del mismo SKU conservan idéntica galería",
      len({tuple(f.imagenes) for f in filas if f.sku == "500"}) == 1)
escribir_xlsx(filas, cfg, salida.name)
wo = openpyxl.load_workbook(salida.name, data_only=True); so = wo.active
encabezados = [c.value for c in so[1]]; idx = {h: i + 1 for i, h in enumerate(encabezados)}
check("plantilla conserva exactamente 36 columnas", encabezados == COLUMNAS and len(encabezados) == 36)
check("CAUPLAS escribe hasta diez imágenes y conserva posición",
      so.cell(2, idx["Imagen1"]).value == "https://ik.imagekit.io/cauplas/500.jpg" and
      so.cell(2, idx["Imagen2"]).value == "https://ik.imagekit.io/cauplas/500-2.jpg" and
      so.cell(2, idx["Imagen3"]).value is None)
check("el Excel conserva las unidades de las medidas en descripción",
      "Largo: 40 cm | 45 cm" in so.cell(2, idx["Descripcion"]).value and
      "Diámetro 1: 32 mm" in so.cell(2, idx["Descripcion"]).value)
check("stock queda exclusivamente en CAUPLAS",
      so.cell(2, idx["Cantidad"]).value == 10 and so.cell(2, idx["CAUPLAS"]).value == 10 and
      all(so.cell(2, idx[b]).value == 0 for b in ("AG", "KG", "KIM", "MATRIZ", "VAZLO")))

print("\n=== 4B. STOCK POR SKU ===")
stocks = openpyxl.Workbook(); ss = stocks.active; ss.append(HEADERS)
ss.append(fila("ST-OK", stock=10.0)); ss.append(fila("ST-OK", fecha="2013", inicio=2013, fin=2013, stock=10))
ss.append(fila("ST-0", stock=0)); ss.append(fila("ST-NEG", stock=-1))
ss.append(fila("ST-DEC", stock=1.5)); ss.append(fila("ST-TXT", stock="mucho"))
ss.append(fila("ST-DIF", stock=2)); ss.append(fila("ST-DIF", fecha="2013", inicio=2013, fin=2013, stock=3))
rs = leer_catalogo_detallado(guardar(stocks), perfil_de("CAUPLAS"))
check("10.0 se interpreta como entero y se repite en cada variante",
      next(p for p in rs.piezas if p["clave"] == "ST-OK")["stock"] == 10 and
      {f.stock for f in generar_filas_con_reporte(rs.piezas, cfg)[0] if f.sku == "ST-OK"} == {10})
check("stock 0 es válido", next(p for p in rs.piezas if p["clave"] == "ST-0")["stock"] == 0)
check("stock inválido o diferente excluye el SKU completo",
      {"ST-NEG", "ST-DEC", "ST-TXT", "ST-DIF"}.isdisjoint({p["clave"] for p in rs.piezas}) and
      {e["clave"] for e in rs.errores} >= {"ST-NEG", "ST-DEC", "ST-TXT", "ST-DIF"}, rs.errores)
check("precio reutiliza la fórmula vigente y activa envío gratis",
      so.cell(2, idx["Precio"]).value == 200 and so.cell(2, idx["Envio Gratis(si,no)"]).value == "No",
      so.cell(2, idx["Precio"]).value)

print("\n=== 5. LÍNEA BASE REAL ===")
real = Path(__file__).resolve().parents[2] / "archivos/publicaciones-masivas/master-cauplas.xlsx"
if real.exists():
    try:
        rr = leer_catalogo_detallado(real, perfil_de("CAUPLAS"))
    except ValueError as exc:
        check("master real histórico sin stock se rechaza", "stock" in str(exc).lower(), str(exc))
    else:
        check("master real: 15,612 filas", rr.filas_master == 15612, rr.filas_master)
        check("master real: 4,214 SKU observados", rr.sku_unicos_master == 4214, rr.sku_unicos_master)
        check("master real: 370 universales", rr.universales == 370, rr.universales)
        check("master real: 15,202 compatibilidades con años", rr.compatibilidades_validas == 15202,
              rr.compatibilidades_validas)
        check("master real: 40 exclusiones por años", rr.compatibilidades_invalidas == 40,
              rr.compatibilidades_invalidas)
else:
    print("  · master real no disponible; se omite la línea base local")

print(f"\n{'=' * 54}\n  {_ok} pasaron · {_fail} fallaron\n{'=' * 54}\n")
sys.exit(1 if _fail else 0)
