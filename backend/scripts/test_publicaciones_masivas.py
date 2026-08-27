"""
Regresión del Módulo 2 (publicaciones masivas).

Fija las trampas que un refactor rompería EN SILENCIO — es decir, las que no
truenan: generan un Excel que se ve bien y se sube a Mercado Libre con datos
equivocados. Correr antes de commitear cambios en aplicaciones_kg.py,
generador_plantilla.py, precio_publicacion.py o parser_catalogo.py.

    /usr/bin/python3 scripts/test_publicaciones_masivas.py
"""
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from services.aplicaciones_kg import parse_aplicaciones
from services.generador_plantilla import (
    COLUMNAS, ConfiguracionProveedor, MAX_TITULO, generar_filas, generar_filas_con_reporte,
    escribir_xlsx,
)
from services.parser_catalogo import (cruzar, cruzar_variantes, leer_catalogo_detallado,
                                      leer_publicaciones)
from services.perfiles_catalogo import perfil_de
import openpyxl
from services.precio_publicacion import ParametrosPrecio, calcular_precio

_ok = _fail = 0


def check(nombre, cond, extra=""):
    global _ok, _fail
    if cond:
        _ok += 1
        print(f"  ✓ {nombre}")
    else:
        _fail += 1
        print(f"  ✗ {nombre}  {extra}")


print("\n=== 1. HERENCIA de marca y modelo ===")
# 🔴 'V8 6.0L 2007-2009' NO es un modelo llamado V8: es el mismo Avalanche con
# otro motor. Sin herencia el título sale 'Bomba de agua P/ V8 6.0' — basura.
r = parse_aplicaciones("GM AVALANCHE V8 5.3L 2007-2010 | V8 6.0L 2007-2009 | SILVERADO 1500 V8 4.8L 2007-2010")
check("un fragmento sin modelo hereda marca Y modelo",
      r.aplicaciones[1].vehiculo == "GM AVALANCHE", r.aplicaciones[1].vehiculo)
check("un modelo nuevo hereda sólo la marca",
      r.aplicaciones[2].vehiculo == "GM SILVERADO 1500", r.aplicaciones[2].vehiculo)

# 🔴 Si el fragmento trae su PROPIA marca, no se le antepone la heredada.
r = parse_aplicaciones("ST CORDOBA L4 1.6L 2004-2009 | IBIZA L4 1.6L 2003-2008 | VW CROSSFOX L4 1.6L 2005-2012")
check("marca propia REEMPLAZA a la heredada (no 'ST VW Crossfox')",
      r.aplicaciones[2].vehiculo == "VW CROSSFOX", r.aplicaciones[2].vehiculo)
check("sin marca propia sí hereda (IBIZA es un Seat)",
      r.aplicaciones[1].vehiculo == "ST IBIZA", r.aplicaciones[1].vehiculo)

# ⚠️ El primer token NO siempre es marca: el catálogo trae códigos de pieza
# ('MANG', 'RAD', 'TA'). Tomarlos por marca metería basura en el título.
r = parse_aplicaciones("GM ASTRA L4 1.8L 2000-2003 | MANG L4 1.4L 2002-2004")
check("un código de pieza NO se toma por marca",
      r.aplicaciones[1].marca == "GM", r.aplicaciones[1].marca)

print("\n=== 2. TRUNCADO del catálogo (el export corta a 90 caracteres) ===")
# 🔴 Sin años no se puede publicar: el título de Gaby SIEMPRE los lleva y una
# compatibilidad sin años no le dice nada al comprador. Publicar una pieza
# diciendo que sirve para menos autos de los que sirve es peor que no publicarla.
r = parse_aplicaciones("CHR 300 V6 3.5L 2005-2010 | PACIFICA V6 3.5L 2005-2006 | SEBRING V6 3.5L")
check("una aplicación SIN años se marca truncada", r.aplicaciones[2].truncada)
check("las completas NO se marcan truncadas",
      not r.aplicaciones[0].truncada and not r.aplicaciones[1].truncada)
check("utiles excluye las truncadas", len(r.utiles) == 2, len(r.utiles))

r = parse_aplicaciones("GM AVALANCHE V8 5.3L 2007-2010 | V8")
check("un resto de 2 letras se marca truncado", r.aplicaciones[1].truncada)

print("\n=== 3. SEPARADORES y formatos del catálogo ===")
check("separa por pipe", len(parse_aplicaciones("FD KA L4 1.3L 1998-2000 | FT PALIO L4 1.6L 2004-2006").utiles) == 2)
check("separa por salto de línea (9 piezas del catálogo)",
      len(parse_aplicaciones("FD KA L4 1.3L 1998-2000\nFT PALIO L4 1.6L 2004-2006").utiles) == 2)
check("una sola aplicación sin separador (1,523 piezas)",
      len(parse_aplicaciones("FT PALIO L4 1.6L 2004-2006").utiles) == 1)
check("celda vacía no truena", len(parse_aplicaciones(None).aplicaciones) == 0)

# Años de 2 dígitos: '98-00' es 1998-2000, no 98 d.C.
a = parse_aplicaciones("FD FIESTA L4 1.3L 98-00").aplicaciones[0]
check("años de 2 dígitos se expanden al siglo correcto",
      (a.anio_desde, a.anio_hasta) == (1998, 2000), (a.anio_desde, a.anio_hasta))
# El catálogo escribe '2.OL' con la letra O en vez del cero.
a = parse_aplicaciones("AUDI TT QUATTRO L4 2.OL TFSI 2015-2018").aplicaciones[0]
check("'2.OL' (letra O) se lee como 2.0L", a.motor == "L4 2.0L", a.motor)

print("\n=== 4. PRECIO: el 13% se DIVIDE, no se suma ===")
# 🔴 La comisión de ML se cobra sobre el precio FINAL. Sumar 13% al costo deja
# la utilidad corta ~$15 por pieza; sobre 3,607 publicaciones no es redondeo.
p = ParametrosPrecio(iva=0.16, utilidad=0.50, comision_ml=0.13, envio_default=100.0)
precio = calcular_precio(346.84, "BOMBA DE AGUA", p)
base = 346.84 * 1.16 * 1.50
check("el precio despeja la comisión (divide)", abs(precio - (base + 100) / 0.87) < 0.02, precio)
check("dividir da MÁS que sumar (la diferencia es la utilidad perdida)",
      precio > base + 100 + base * 0.13)
comision_real = precio * 0.13
check("tras pagar comisión y envío queda la utilidad completa",
      abs((precio - comision_real - 100) - base) < 0.02)

# None y 0.0 son distintos: un 0.0 se publicaría en ML como precio real.
check("costo no numérico -> None (celda vacía), NUNCA 0.0", calcular_precio("n/a") is None)
check("costo 0 -> None", calcular_precio(0) is None)
check("costo negativo -> None", calcular_precio(-5) is None)
check("comisión del 100% no inventa precio",
      calcular_precio(100, "", ParametrosPrecio(comision_ml=1.0)) is None)

# ⬜ El envío está PENDIENTE: por default no suma. Que el default sea 0 y no un
# valor inventado es deliberado — un envío falso descuadra el precio en silencio.
check("sin tabla de envío el default es 0 (hueco visible, no inventado)",
      calcular_precio(100.0, "RADIADOR", ParametrosPrecio()) == round(100 * 1.16 * 1.5 / 0.87, 2))

print("\n=== 5. EXPANSIÓN: una pieza -> N publicaciones ===")
cfg = ConfiguracionProveedor(codigo_bodega="KG", marca="KeepOnGreen", categoria_ml="MLM163963")
piezas = [{
    "clave": "KGP-1106", "linea": "BOMBA DE AGUA", "costo": 500.0,
    "aplicaciones": "GM AVALANCHE V8 5.3L 2007-2010 | V8 6.0L 2007-2009 | SILVERADO 1500 V8 4.8L 2007-2010",
}]
filas = generar_filas(piezas, cfg)
check("3 aplicaciones -> 3 publicaciones", len(filas) == 3, len(filas))
check("todas comparten el MISMO SKU", len({f.sku for f in filas}) == 1)
check("todas comparten el MISMO precio", len({f.precio for f in filas}) == 1)
check("cada una tiene título DISTINTO", len({f.titulo for f in filas}) == 3)

# Las truncadas no generan publicación salvo que se pidan explícitamente.
piezas_t = [{"clave": "K1", "linea": "BOMBA DE AGUA", "costo": 100.0,
             "aplicaciones": "GM AVALANCHE V8 5.3L 2007-2010 | V8"}]
check("las truncadas NO se publican por default", len(generar_filas(piezas_t, cfg)) == 1)
check("incluir_truncadas=True sí las trae (para revisión)",
      len(generar_filas(piezas_t, cfg, incluir_truncadas=True)) == 2)

print("\n=== 6. TÍTULO: tope de 60 caracteres de Mercado Libre ===")
largo = [{"clave": "K", "linea": "KIT BANDA DISTRIBUCION C/BOMBA", "costo": 100.0,
          "aplicaciones": "VOLKSWAGEN CROSSFOX SPORTLINE L4 1.6L 2005-2018"}]
f = generar_filas(largo, cfg)[0]
check(f"título recortado a {MAX_TITULO}", len(f.titulo) <= MAX_TITULO, f"{len(f.titulo)}: {f.titulo}")
check("no se corta a media palabra (suelta años del final)", not f.titulo.endswith(("-", " 2")), f.titulo)

print("\n=== 7. DESCRIPCIÓN: cabeza variable + cuerpo fijo ===")
cfg_desc = ConfiguracionProveedor(codigo_bodega="KG", descripcion_base="GARANTÍA RELUVSA\nHorarios: L-V")
f = generar_filas(piezas, cfg_desc)[0]
check("lleva el nombre de la pieza", "BOMBA DE AGUA" in f.descripcion)
check("lleva el OEM (la clave del proveedor)", "KGP-1106" in f.descripcion)
check("lleva la sección Compatibilidades", "Compatibilidades:" in f.descripcion)
check("lleva el cuerpo fijo del proveedor", "GARANTÍA RELUVSA" in f.descripcion)
check("las compatibilidades listan TODAS las aplicaciones, no sólo la de su fila",
      "SILVERADO" in f.descripcion and "AVALANCHE" in f.descripcion)
# Las hermanas comparten descripción: es la misma pieza.
check("todas las filas de una pieza comparten descripción",
      len({x.descripcion for x in generar_filas(piezas, cfg_desc)}) == 1)

print("\n=== 8. PLANTILLA: las 36 columnas de Mercado Libre ===")
check("son exactamente 36 columnas", len(COLUMNAS) == 36, len(COLUMNAS))
check("no hay columnas repetidas", len(set(COLUMNAS)) == 36)
check("el orden arranca con Titulo/Categoria/Precio",
      COLUMNAS[:3] == ["Titulo", "Categoria", "Precio"])
check("las 6 bodegas van al final",
      COLUMNAS[-6:] == ["AG", "CAUPLAS", "KG", "KIM", "MATRIZ", "VAZLO"])
check("Imagen1..10 existen (van vacías, Gaby las pega)",
      all(f"Imagen{i}" in COLUMNAS for i in range(1, 11)))

print("\n=== 9. CRUCE contra lo ya publicado ===")
# ⚠️ La columna Q trae paquetes con '&': sin partirla, un SKU publicado dentro
# de un paquete se contaría como faltante.
cat = [{"clave": "KGP-1449"}, {"clave": "KGP-9999"}, {"clave": "NO625HB"}]
res = cruzar(cat, {"KGP-1449", "NO625HB"})
check("separa publicadas de faltantes", (res["ya_publicadas"], res["faltantes"]) == (2, 1))
check("el total cuadra", res["total_catalogo"] == 3)
check("el cruce ignora mayúsculas", cruzar([{"clave": "kgp-1449"}], {"KGP-1449"})["ya_publicadas"] == 1)

print("\n=== 10. NUEVO MASTER KG: validación y consolidación ===")
headers = ["Armadora", "Modelo", "Motor", "Producto", "Año", "Inicio", "Fin", "Clave",
           "Especificaciones", "Caracteristicas", "Guia de Compradores", "OEM"] + [f"Imagen {i}" for i in range(1, 6)]
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "BD_Catalogo"; ws.append(headers)
base = ["ACURA", "CL", "L4 2.2L", "Bomba  de Agua", 1997, 1997, 1997, "KGP-759",
        "Aluminio", "Impulsor", "ACURA CL L4 2.2L 1997", "19200P0A003"] + [f"img{i}" for i in range(1, 6)]
ws.append(base); ws.append(base)
ws.append(["ACURA", "CL", "L4 2.3L", "bomba de agua", "1998-1999", 1998, 1999, "KGP-759",
           "Aluminio", "Impulsor", "ACURA CL L4 2.3L 1998-1999", "ALT-002"] + [f"img{i}" for i in range(1, 6)])
ws.append(["VW", "POLO", "L4 1.6L", "Toma de Agua", "2003-2001", 2003, 2001, "BAD",
           "", "", "VW POLO", "-"] + [""] * 5)
tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); tmp.close(); wb.save(tmp.name)
r = leer_catalogo_detallado(tmp.name, perfil_de("KG")); p = r.piezas[0]
check("detecta el master en BD_Catalogo", r.formato == "master_kg")
check("agrupa compatibilidades por Clave", len(r.piezas) == 1 and len(p["compatibilidades"]) == 2)
check("descarta fila exactamente duplicada", r.duplicados_descartados == 1)
check("excluye Inicio > Fin con fila y motivo", r.compatibilidades_invalidas == 1 and r.errores[0]["fila"] == 5)
check("normaliza Producto y consolida OEM únicos", p["producto"] == "Bomba de Agua" and p["oems"] == ["19200P0A003", "ALT-002"])
check("sin Precio se permite y queda costo vacío", not r.precio_presente and p["costo"] is None)

# El nombre de la hoja no define el formato. También se normalizan espacios,
# mayúsculas y acentos de los encabezados antes de reconocer la estructura.
ws.title = "Catálogo vigente agosto"
ws.cell(1, 1).value = "  ARMADORA "
ws.cell(1, 4).value = " producto "
ws.cell(1, 5).value = "ANO"
renombrado = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); renombrado.close()
wb.save(renombrado.name)
rr = leer_catalogo_detallado(renombrado.name, perfil_de("KG"))
check("detecta master por encabezados aunque la hoja cambie de nombre", rr.formato == "master_kg")
check("la categoría sale de Producto, nunca de Modelo",
      {x["linea"] for x in rr.piezas} == {"Bomba de Agua"})

incompleto = openpyxl.Workbook(); wi = incompleto.active; wi.title = "Carga KG"
wi.append(["Armadora", "Modelo", "Producto", "Clave", "Inicio"])
ti = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); ti.close(); incompleto.save(ti.name)
try:
    leer_catalogo_detallado(ti.name, perfil_de("KG"))
    error_incompleto = ""
except ValueError as exc:
    error_incompleto = str(exc)
check("un master incompleto no cae al catálogo legado",
      "parece ser el master KG" in error_incompleto and "Fin" in error_incompleto,
      error_incompleto)

print("\n=== 11. VARIANTES, DESCRIPCIÓN E IMÁGENES DEL MASTER ===")
filas_n, reporte_n = generar_filas_con_reporte(r.piezas, cfg_desc)
check("un año genera una sola variante", sum("1997" in f.titulo for f in filas_n) == 1)
check("dos años generan rango y dos bloques", sum(f.sku == "KGP-759" for f in filas_n) == 4, [f.titulo for f in filas_n])
check("ningún título lleva OEM ni Clave", all("19200" not in f.titulo and "KGP-759" not in f.titulo for f in filas_n))
check("todos los títulos respetan 60", all(len(f.titulo) <= 60 for f in filas_n))
check("todos los OEM van en descripción", all("OEM: 19200P0A003, ALT-002" in f.descripcion for f in filas_n))
check("Imagen 1-5 se conserva", all(f.imagenes == [f"img{i}" for i in range(1, 6)] for f in filas_n))

larga = [{"clave": "L1", "linea": "Kit de Banda de Distribución c/bomba",
          "producto": "Kit de Banda de Distribución c/bomba", "formato": "master_kg", "costo": None,
          "oems": [], "especificaciones": [], "caracteristicas": [], "imagenes": [""]*5,
          "compatibilidades": [{"armadora": "VOLKSWAGEN", "modelo": "CROSSFOX SPORTLINE",
          "motor": "L4 1.6L", "inicio": 2005, "fin": 2018, "guia": "", "fila": 2}]}]
lf, lr = generar_filas_con_reporte(larga, cfg)
check("rango largo produce rango y bloques/subbloques", len(lf) >= 3)
check("alias/remoción de armadora evita cortes", all(len(x.titulo) <= 60 and not x.titulo.endswith("-") for x in lf))
imposible = [dict(larga[0], clave="L2", producto="Producto Extraordinariamente Largo Sin Alias",
                  linea="Producto Extraordinariamente Largo Sin Alias",
                  compatibilidades=[dict(larga[0]["compatibilidades"][0], modelo="MODELO EXTRAORDINARIAMENTE LARGO")])]
ix, ir = generar_filas_con_reporte(imposible, cfg)
check("si no cabe sin cortar datos se excluye con motivo", not ix and ir["exclusiones"])

wbp = openpyxl.Workbook(); wsp = wbp.active; wsp.title = "BD_Catalogo"; wsp.append(headers + ["  GRAN   MAYOREO "])
wsp.append(base + [100]); wsp.append([*base[:2], "L4 2.3L", *base[3:5], 1998, 1999, *base[7:], 100])
tpp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); tpp.close(); wbp.save(tpp.name)
rp = leer_catalogo_detallado(tpp.name, perfil_de("KG"))
check("Gran Mayoreo en R se lee como costo sin IVA",
      rp.precio_presente and rp.piezas[0]["costo"] == 100.0)
filas_precio, _ = generar_filas_con_reporte(rp.piezas, cfg_desc)
out_precio = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); out_precio.close()
escribir_xlsx(filas_precio, cfg_desc, out_precio.name)
wb_precio = openpyxl.load_workbook(out_precio.name, data_only=True)
check("Gran Mayoreo genera una celda Precio no vacía con la fórmula vigente",
      wb_precio.active.cell(2, 3).value is not None)

wb_anterior = openpyxl.Workbook(); ws_anterior = wb_anterior.active; ws_anterior.append(headers + ["Precio"])
ws_anterior.append(base + [125])
t_anterior = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); t_anterior.close(); wb_anterior.save(t_anterior.name)
r_anterior = leer_catalogo_detallado(t_anterior.name, perfil_de("KG"))
check("el master anterior con Precio en R continúa funcionando",
      r_anterior.precio_presente and r_anterior.piezas[0]["costo"] == 125.0)

wb_fuera = openpyxl.Workbook(); ws_fuera = wb_fuera.active
ws_fuera.append(["Gran Mayoreo"] + headers); ws_fuera.append([100] + base)
t_fuera = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); t_fuera.close(); wb_fuera.save(t_fuera.name)
try:
    leer_catalogo_detallado(t_fuera.name, perfil_de("KG"))
    error_fuera = ""
except ValueError as exc:
    error_fuera = str(exc)
check("Gran Mayoreo fuera de R se rechaza", "inmediatamente después de Imagen 5" in error_fuera,
      error_fuera)

wsp.cell(3, 18).value = 120
wbp.save(tpp.name)
rp = leer_catalogo_detallado(tpp.name, perfil_de("KG"))
check("precios distintos del SKU no se eligen arbitrariamente",
      rp.precio_presente and rp.sku_precio_inconsistente == 1 and rp.piezas[0]["costo"] is None)

print("\n=== 12. CRUCE SKU + TÍTULO Y PLANTILLA ===")
pub = openpyxl.Workbook(); ps = pub.active; ps.append([""] * 17)
row = [""] * 17; row[1] = filas_n[0].titulo.upper(); row[16] = "OTRO&KGP-759"; ps.append(row)
pt = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); pt.close(); pub.save(pt.name)
pares = leer_publicaciones(pt.name); cv = cruzar_variantes(filas_n, pares)
check("paquete con & asocia el título a cada SKU", ("KGP-759", filas_n[0].titulo.lower()) in pares)
check("cruce excluye sólo SKU+título existente", len(cv["existentes"]) == 1 and len(cv["pendientes"]) == 3)
out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False); out.close(); escribir_xlsx(filas_n, cfg_desc, out.name)
ow = openpyxl.load_workbook(out.name, data_only=True); os = ow.active
check("xlsx copia Imagen1-5 y deja Imagen6 vacía", os.cell(2, 15).value == "img1" and os.cell(2, 19).value == "img5" and os.cell(2, 20).value is None)
check("precio ausente queda celda vacía", os.cell(2, 3).value is None)

print(f"\n{'='*54}")
print(f"  {_ok} pasaron · {_fail} fallaron")
print(f"{'='*54}\n")
sys.exit(1 if _fail else 0)
